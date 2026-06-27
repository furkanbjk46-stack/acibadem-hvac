from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

KOYU_MAVI = '1F3864'
ACIK_MAVI = 'D6E4F0'
SARI      = 'FFF2CC'
YESIL     = 'E2EFDA'
TURUNCU   = 'FCE4D6'
BEYAZ     = 'FFFFFF'
GRI       = 'F2F2F2'

def baslik(ws, cell, text, font_size=12):
    ws[cell] = text
    ws[cell].fill = PatternFill('solid', start_color=KOYU_MAVI)
    ws[cell].font = Font(bold=True, color=BEYAZ, size=font_size, name='Arial')
    ws[cell].alignment = Alignment(horizontal='center', vertical='center')

def alan_baslik(ws, cell, text):
    ws[cell] = text
    ws[cell].fill = PatternFill('solid', start_color=ACIK_MAVI)
    ws[cell].font = Font(bold=True, color=KOYU_MAVI, size=10, name='Arial')
    ws[cell].alignment = Alignment(horizontal='left', vertical='center')

def satir(ws, row, etiket, aciklama='', ornek='', zorunlu=True):
    prefix = '* ' if zorunlu else ''
    ws.cell(row=row, column=1).value = prefix + etiket
    ws.cell(row=row, column=1).fill = PatternFill('solid', start_color=GRI)
    ws.cell(row=row, column=1).font = Font(bold=True, size=9, name='Arial', color='333333')
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')

    ws.cell(row=row, column=2).fill = PatternFill('solid', start_color=SARI)
    ws.cell(row=row, column=2).font = Font(size=10, name='Arial')
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')

    ws.cell(row=row, column=3).value = aciklama
    ws.cell(row=row, column=3).fill = PatternFill('solid', start_color=BEYAZ)
    ws.cell(row=row, column=3).font = Font(size=9, name='Arial', color='666666', italic=True)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    if ornek:
        ws.cell(row=row, column=4).value = 'Ornek: ' + ornek
        ws.cell(row=row, column=4).font = Font(size=9, name='Arial', color='0070C0', italic=True)
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='left', vertical='center')

    ws.row_dimensions[row].height = 20

# ════════════════════════════════════════════════
# SAYFA 1 — LOKASYON PROFILi
# ════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Lokasyon Profili'
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 38
ws1.column_dimensions['D'].width = 26

ws1.row_dimensions[1].height = 34
ws1.merge_cells('A1:D1')
baslik(ws1, 'A1', 'YENi LOKASYON KURULUM SABLONU  —  Lokasyon Profili', font_size=13)

ws1.row_dimensions[2].height = 18
ws1.merge_cells('A2:D2')
ws1['A2'] = 'Sari hucreler doldurulacak alanlardir  |  * isareti zorunlu alandir'
ws1['A2'].fill = PatternFill('solid', start_color=TURUNCU)
ws1['A2'].font = Font(bold=True, size=9, name='Arial', color='833C00')
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')

ws1.row_dimensions[3].height = 8

# Bolum 1
ws1.row_dimensions[4].height = 22
ws1.merge_cells('A4:D4')
alan_baslik(ws1, 'A4', '  1. TEMEL KIMLIK BILGILERI')

satir(ws1, 5,  'Lokasyon ID',           'Sistem ici kisa kod (kucuk harf, bosluksuz)',         'atasehir')
satir(ws1, 6,  'Lokasyon Adi',          'Portalde gorunecek tam isim',                          'Acibadem Atasehir Hastanesi')
satir(ws1, 7,  'Sehir / Ilce',          '',                                                     'Istanbul / Atasehir',       zorunlu=False)
satir(ws1, 8,  'Sorumlu Kisi',          'Teknik irtibat',                                       'Ahmet Yilmaz',             zorunlu=False)
satir(ws1, 9,  'PC IP Adresi',          'Lokasyon PC ag adresi',                                '192.168.1.100',            zorunlu=False)

ws1.row_dimensions[10].height = 8
ws1.row_dimensions[11].height = 22
ws1.merge_cells('A11:D11')
alan_baslik(ws1, 'A11', '  2. SiSTEM YAPISI')

satir(ws1, 12, 'Hat Sayisi',                '1 veya 2',                                         '1')
satir(ws1, 13, 'Hat 1 Adi',                 'Birinci hatin kisa adi',                           'ATS-1')
satir(ws1, 14, 'Hat 2 Adi',                 'Ikinci hat varsa — yoksa bos birakin',             'ATS-2',  zorunlu=False)
satir(ws1, 15, 'Chiller Sayisi',             'Toplam chiller adedi',                            '3')
satir(ws1, 16, 'Absorpsiyonlu Chiller',      'Var mi? (EVET / HAYIR)',                          'HAYIR')
satir(ws1, 17, 'Kazan Sayisi',               'Toplam kazan adedi (yoksa 0)',                    '2')
satir(ws1, 18, 'Kule (Cooling Tower)',        'Toplam kule adedi (yoksa 0)',                    '2')
satir(ws1, 19, 'VRF / Split Sistemi',        'Var mi? (EVET / HAYIR)',                          'HAYIR')
satir(ws1, 20, 'Kojenerasyon',               'Var mi? (EVET / HAYIR)',                          'HAYIR')

ws1.row_dimensions[21].height = 8
ws1.row_dimensions[22].height = 22
ws1.merge_cells('A22:D22')
alan_baslik(ws1, 'A22', '  3. SAYAC / OLCUM ALTYAPISI')

satir(ws1, 23, 'Modbus Analizor Var mi',    'Janitza/Siemens sayac (EVET/HAYIR)',               'EVET')
satir(ws1, 24, 'BACnet (Desigo) Var mi',    'Desigo CC BACnet okuma (EVET/HAYIR)',              'EVET')
satir(ws1, 25, 'Modbus Port',               'Standart 502 — farkliysa yaz',                    '502',    zorunlu=False)
satir(ws1, 26, 'BACnet Gateway IP',         'BACnet/IP gateway adresi',                        '172.17.91.50', zorunlu=False)
satir(ws1, 27, 'Veri Okuma Saati',          'Bilgi amaclidir — sistemde SABIT 07:10 (data_bridge.py), buraya yazilan deger otomatik uygulanmaz', '07:10', zorunlu=False)
satir(ws1, 28, 'GM Sync Saati',             'Bilgi amaclidir — sistemde SABIT 08:00 (cloud_sync.py), buraya yazilan deger otomatik uygulanmaz', '08:00', zorunlu=False)

ws1.row_dimensions[29].height = 8
ws1.row_dimensions[30].height = 22
ws1.merge_cells('A30:D30')
alan_baslik(ws1, 'A30', '  4. MANUEL GiRiLECEK SAYACLAR (Portalde form olacak)')

satir(ws1, 31, 'Sebeke Tuketim Sayaci',     'kWh sayaci var mi? (EVET/HAYIR)',                  'EVET')
satir(ws1, 32, 'Kojenerasyon Uretim',       'kWh uretim sayaci? (EVET/HAYIR)',                  'HAYIR')
satir(ws1, 33, 'Dogalgaz Sayaci (Kazan)',   'm3 sayaci? (EVET/HAYIR)',                          'EVET')
satir(ws1, 34, 'Dogalgaz (Kojenerasyon)',   'm3 sayaci? (EVET/HAYIR)',                          'HAYIR')
satir(ws1, 35, 'Su Tuketim Sayaci',         'm3 sayaci? (EVET/HAYIR)',                          'EVET')

ws1.row_dimensions[36].height = 8
ws1.row_dimensions[37].height = 22
ws1.merge_cells('A37:D37')
alan_baslik(ws1, 'A37', '  5. GENEL AHU UFLEME BANDI (Opsiyonel — bos ise varsayilan kullanilir)')

satir(ws1, 38, 'Sogutma Min Ufleme (SAT)',  'Bos = varsayilan 15.0 kullanilir',                 '15.0', zorunlu=False)
satir(ws1, 39, 'Sogutma Max Ufleme (SAT)',  'Bos = varsayilan 18.0 kullanilir',                 '18.0', zorunlu=False)
satir(ws1, 40, 'Isitma Min Ufleme (SAT)',   'Bos = varsayilan 27.0 kullanilir',                 '27.0', zorunlu=False)
satir(ws1, 41, 'Isitma Max Ufleme (SAT)',   'Bos = varsayilan 30.0 kullanilir',                 '30.0', zorunlu=False)

ws1.row_dimensions[42].height = 8
ws1.row_dimensions[43].height = 22
ws1.merge_cells('A43:D43')
alan_baslik(ws1, 'A43', '  6. CHILLER / FCU KAPASITE BILGISI (Opsiyonel — kojen/uretim hesabi icin)')

satir(ws1, 44, 'Chiller Birim Kapasite (kW)', 'Tek bir chiller biriminin kW kapasitesi',         '2000', zorunlu=False)
satir(ws1, 45, 'Chiller Gidis Sicakligi (C)', 'Tasarim gidis suyu sicakligi',                    '7',    zorunlu=False)
satir(ws1, 46, 'Chiller Donus Sicakligi (C)', 'Tasarim donus suyu sicakligi',                    '12',   zorunlu=False)
satir(ws1, 47, 'FCU Adedi',                   'Toplam FCU sayisi',                               '1750', zorunlu=False)
satir(ws1, 48, 'FCU Birim Kapasite (kW)',     'Tek bir FCU biriminin ortalama kW kapasitesi',     '2.5',  zorunlu=False)
satir(ws1, 49, 'FCU Esanjor Diversity',       'Eszamanlilik faktoru (0-1 arasi)',                 '0.55', zorunlu=False)
satir(ws1, 50, 'Tasarim Hava DT (C)',         'AHU/FCU tasarim hava sicaklik farki',              '5',    zorunlu=False)

ws1.row_dimensions[51].height = 8
ws1.row_dimensions[52].height = 22
ws1.merge_cells('A52:D52')
alan_baslik(ws1, 'A52', '  7. LiSANS BiLGiSi (Lokasyon ID ile KARISTIRMAYIN — farkli kavramlar)')

satir(ws1, 53, 'Makine ID',  'Lokasyon PC sinin anakart UUID si — PC de "wmic csproduct get uuid" '
                              'komutuyla alinir. Bu, Lokasyon ID DEGiLDiR, ayri bir kavramdir.',
      'DF9ABA0D-A7A8-C416-A430-00D861BD5794', zorunlu=False)

# ════════════════════════════════════════════════
# SAYFA 2 — MODBUS ANALIZORLER
# ════════════════════════════════════════════════
ws2 = wb.create_sheet('Modbus Analizorler')
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 16   # Kategori (YENi)
ws2.column_dimensions['F'].width = 28   # Notlar

ws2.row_dimensions[1].height = 30
ws2.merge_cells('A1:F1')
baslik(ws2, 'A1', 'MODBUS ANALIZOR LiSTESi')

ws2.row_dimensions[2].height = 18
ws2.merge_cells('A2:F2')
ws2['A2'] = 'Her satir bir analizoru temsil eder. Marka: janitza / siemens  |  Kategori: Chiller / MCC / Kule / TRDP / Diger'
ws2['A2'].fill = PatternFill('solid', start_color=ACIK_MAVI)
ws2['A2'].font = Font(size=9, name='Arial', color=KOYU_MAVI, italic=True)
ws2['A2'].alignment = Alignment(horizontal='left', vertical='center')

ws2.row_dimensions[3].height = 22
for col, text in enumerate(['#', 'Cihaz Adi', 'IP Adresi', 'Marka', 'Kategori', 'Notlar'], 1):
    c = ws2.cell(row=3, column=col)
    c.value = text
    c.fill = PatternFill('solid', start_color=KOYU_MAVI)
    c.font = Font(bold=True, color=BEYAZ, size=10, name='Arial')
    c.alignment = Alignment(horizontal='center', vertical='center')

for i in range(4, 30):
    ws2.row_dimensions[i].height = 20
    ws2.cell(row=i, column=1).value = i - 3
    for col in range(1, 7):
        ws2.cell(row=i, column=col).fill = PatternFill('solid', start_color=SARI)
        ws2.cell(row=i, column=col).font = Font(size=10, name='Arial')
        ws2.cell(row=i, column=col).alignment = Alignment(horizontal='left', vertical='center')

# Ornek satir (4. satir) - dogru isimlendirme formatini gostermek icin
ws2.cell(row=4, column=2).value = 'MCC-1'
ws2.cell(row=4, column=3).value = '172.17.91.100'
ws2.cell(row=4, column=4).value = 'janitza'
ws2.cell(row=4, column=5).value = 'MCC'
ws2.cell(row=4, column=6).value = 'ORNEK SATIR - silip gercek degerleri girin'
for col in range(2, 7):
    ws2.cell(row=4, column=col).font = Font(size=10, name='Arial', italic=True, color='0070C0')

# Marka dropdown: D4:D29
dv_marka = DataValidation(type='list', formula1='"janitza,siemens"', showDropDown=False)
dv_marka.sqref = 'D4:D29'
ws2.add_data_validation(dv_marka)

# Kategori dropdown: E4:E29
dv_kategori = DataValidation(type='list', formula1='"Chiller,MCC,Kule,TRDP,Diger"', showDropDown=False)
dv_kategori.sqref = 'E4:E29'
ws2.add_data_validation(dv_kategori)

ws2.row_dimensions[30].height = 18
ws2.merge_cells('A30:F30')
ws2['A30'] = 'Kategori:  Chiller = sogutma gurubu  |  MCC = elektrik panosu  |  Kule = sogutma kulesi  |  TRDP = trafo dagitim panosu  |  Diger = siniflandirilmamis'
ws2['A30'].fill = PatternFill('solid', start_color=YESIL)
ws2['A30'].font = Font(size=9, name='Arial', color='375623')
ws2['A30'].alignment = Alignment(horizontal='left', vertical='center')

ws2.row_dimensions[31].height = 36
ws2.merge_cells('A31:F31')
ws2['A31'] = ('ONEMLI — Cihaz Adi formati: Sistem cihazlari BUYUK HARF ve tire ile esler (case-sensitive). '
              'Ornekler: MCC-1, MCC-2 ... MCC-7  |  CHILLER-1 ... CHILLER-5  |  KULE-1, KULE-2, KULE-3  |  '
              'TRDP-1, TRDP-3  |  Siemens MCC panelleri icin: 2BK-MCC-D01, 2BK-MCC-D02, 4BK-MCC-E01, 4BK-MCC-E02, '
              '4BK-MCC-F01, CK-MCC-D01, CK-MCC-E01, CK-MCC-F01. Bu formata uymayan isimler (orn. "MCC1" veya "mcc-1") '
              'sistemde MCC/Chiller/Kule kirilim raporlarinda eslesmez ve o cihazin verisi ayri gosterilemez.')
ws2['A31'].fill = PatternFill('solid', start_color=TURUNCU)
ws2['A31'].font = Font(size=9, name='Arial', color='833C00', italic=True)
ws2['A31'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ════════════════════════════════════════════════
# SAYFA 3 — BACNET NOKTALAR
# ════════════════════════════════════════════════
ws3 = wb.create_sheet('BACnet Noktalar')
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 13
ws3.column_dimensions['F'].width = 16
ws3.column_dimensions['G'].width = 28
ws3.column_dimensions['H'].width = 22
ws3.column_dimensions['I'].width = 30   # Sistem Rolu (YENi)

ws3.row_dimensions[1].height = 30
ws3.merge_cells('A1:I1')
baslik(ws3, 'A1', 'BACNET NOKTA LiSTESi (Desigo CC)')

ws3.row_dimensions[2].height = 18
ws3.merge_cells('A2:I2')
ws3['A2'] = 'Sistem Rolu sutunu: bu noktanin sistemde hangi degeri temsil ettigini belirtir (acilir liste)'
ws3['A2'].fill = PatternFill('solid', start_color=ACIK_MAVI)
ws3['A2'].font = Font(size=9, name='Arial', color=KOYU_MAVI, italic=True)
ws3['A2'].alignment = Alignment(horizontal='left', vertical='center')

ws3.row_dimensions[3].height = 22
for col, text in enumerate(['Gateway IP', 'Network (DNET)', 'MAC (DADR)',
                              'Device Instance ID', 'Object Type', 'Object Instance',
                              'Point Name', 'Aciklama', 'Sistem Rolu'], 1):
    c = ws3.cell(row=3, column=col)
    c.value = text
    c.fill = PatternFill('solid', start_color=KOYU_MAVI if col != 9 else '375623')
    c.font = Font(bold=True, color=BEYAZ, size=10, name='Arial')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for i in range(4, 55):
    ws3.row_dimensions[i].height = 20
    for col in range(1, 10):
        ws3.cell(row=i, column=col).fill = PatternFill('solid', start_color=SARI if col < 9 else YESIL)
        ws3.cell(row=i, column=col).font = Font(size=10, name='Arial')
        ws3.cell(row=i, column=col).alignment = Alignment(horizontal='left', vertical='center')

# Sistem Rolu dropdown: I4:I54
roller = (
    'Dis Hava Sicakligi,'
    'Chiller Set Sicakligi,'
    'Hat1 Isitma Sicakligi,'
    'Hat1 Sogutma Sicakligi,'
    'Hat1 Kazan Sicakligi,'
    'Hat2 Isitma Sicakligi,'
    'Hat2 Sogutma Sicakligi,'
    'Hat2 Kazan Sicakligi,'
    'Chiller-1 Durum,'
    'Chiller-2 Durum,'
    'Chiller-3 Durum,'
    'Chiller-4 Durum,'
    'Chiller-5 Durum,'
    'Chiller-1 Yuzde,'
    'Chiller-2 Yuzde,'
    'Chiller-3 Yuzde,'
    'Chiller-4 Yuzde,'
    'Chiller-5 Yuzde,'
    'Absorpsiyon Chiller Durum,'
    'Kazan-1 Durum,'
    'Kazan-2 Durum,'
    'Kazan-3 Durum,'
    'Kullanilmayacak'
)
dv_rol = DataValidation(type='list', formula1=f'"{roller}"', showDropDown=False)
dv_rol.sqref = 'I4:I54'
ws3.add_data_validation(dv_rol)

# Aciklama satiri
ws3.row_dimensions[55].height = 18
ws3.merge_cells('A55:I55')
ws3['A55'] = 'Sistem Rolu secilmeyen (bos birakilan) noktalar okunur ama energy_data sutunlarina yazilmaz'
ws3['A55'].fill = PatternFill('solid', start_color=YESIL)
ws3['A55'].font = Font(size=9, name='Arial', color='375623', italic=True)
ws3['A55'].alignment = Alignment(horizontal='left', vertical='center')

# ════════════════════════════════════════════════
# SAYFA 4 — AHU NOKTALARI (ahu_collector.py --setup Book2 formati)
# ════════════════════════════════════════════════
ws_ahu = wb.create_sheet('AHU Noktalari')
ws_ahu.column_dimensions['A'].width = 18
ws_ahu.column_dimensions['B'].width = 15
ws_ahu.column_dimensions['C'].width = 14
ws_ahu.column_dimensions['D'].width = 20
ws_ahu.column_dimensions['E'].width = 13
ws_ahu.column_dimensions['F'].width = 16
ws_ahu.column_dimensions['G'].width = 28
ws_ahu.column_dimensions['H'].width = 20

ws_ahu.row_dimensions[1].height = 30
ws_ahu.merge_cells('A1:H1')
baslik(ws_ahu, 'A1', 'AHU NOKTA LiSTESi (Mekanik Zeka / HVAC Analiz)')

ws_ahu.row_dimensions[2].height = 30
ws_ahu.merge_cells('A2:H2')
ws_ahu['A2'] = ('Bu sayfa SANTRAL (chiller/kazan) noktalarindan AYRIDIR — sadece AHU (klima santrali) '
                'noktalari icin kullanilir. ahu_collector.py --setup ile islenir.')
ws_ahu['A2'].fill = PatternFill('solid', start_color=ACIK_MAVI)
ws_ahu['A2'].font = Font(size=9, name='Arial', color=KOYU_MAVI, italic=True)
ws_ahu['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws_ahu.row_dimensions[3].height = 22
for col, text in enumerate(['Gateway IP', 'Network (DNET)', 'MAC (DADR)',
                              'Device Instance ID', 'Object Type', 'Object Instance',
                              'Point Name', 'LOCATION (MAHAL)'], 1):
    c = ws_ahu.cell(row=3, column=col)
    c.value = text
    c.fill = PatternFill('solid', start_color=KOYU_MAVI)
    c.font = Font(bold=True, color=BEYAZ, size=10, name='Arial')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for i in range(4, 55):
    ws_ahu.row_dimensions[i].height = 20
    for col in range(1, 9):
        ws_ahu.cell(row=i, column=col).fill = PatternFill('solid', start_color=SARI)
        ws_ahu.cell(row=i, column=col).font = Font(size=10, name='Arial')
        ws_ahu.cell(row=i, column=col).alignment = Alignment(horizontal='left', vertical='center')

# Ornek satir (4. satir)
ws_ahu.cell(row=4, column=1).value = '172.17.91.50'
ws_ahu.cell(row=4, column=2).value = 2
ws_ahu.cell(row=4, column=3).value = '0x011C'
ws_ahu.cell(row=4, column=4).value = 2098211
ws_ahu.cell(row=4, column=5).value = 0
ws_ahu.cell(row=4, column=6).value = 40
ws_ahu.cell(row=4, column=7).value = 'Ahu-3 Ufleme'
ws_ahu.cell(row=4, column=8).value = 'MAS-1'
for col in range(1, 9):
    ws_ahu.cell(row=4, column=col).font = Font(size=10, name='Arial', italic=True, color='0070C0')

ws_ahu.row_dimensions[55].height = 54
ws_ahu.merge_cells('A55:H55')
ws_ahu['A55'] = (
    'ONEMLI — Point Name formati: ILK KELIME AHU adidir (orn. "Ahu-3"), sistem bunu otomatik ayirir. '
    'Devamindaki kelimeye gore nokta tipi otomatik tespit edilir: '
    '"Ufleme" iceren -> SAT/ufleme sicakligi  |  "...Sogutma Vana..." -> sogutma vana yuzdesi  |  '
    '"...Isitma Vana..." -> isitma vana yuzdesi  |  "...Set" ile bitenler -> set sicakligi  |  '
    'digerleri -> emis/donus sicakligi. Ornekler: "Ahu-3 Ufleme", "Ahu-3 Emis", "Ahu-3 Isitma Vana", '
    '"Ahu-3 Sogutma Vana", "Ahu-3 Set". LOCATION (MAHAL) sutunu hangi hat/bolgede oldugunu belirtir '
    '(orn. MAS-1, MAS-2 — Lokasyon Profili sayfasindaki Hat 1/Hat 2 adlarina karsilik gelir).'
)
ws_ahu['A55'].fill = PatternFill('solid', start_color=TURUNCU)
ws_ahu['A55'].font = Font(size=9, name='Arial', color='833C00', italic=True)
ws_ahu['A55'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ════════════════════════════════════════════════
# SAYFA 5 — AHU SAT LIMITLERI / KAPASITE
# ════════════════════════════════════════════════
ws_sat = wb.create_sheet('AHU SAT ve Kapasite')
ws_sat.column_dimensions['A'].width = 14
ws_sat.column_dimensions['B'].width = 14
ws_sat.column_dimensions['C'].width = 22
ws_sat.column_dimensions['D'].width = 22
ws_sat.column_dimensions['E'].width = 22

ws_sat.row_dimensions[1].height = 30
ws_sat.merge_cells('A1:E1')
baslik(ws_sat, 'A1', 'AHU SAT LiMiTLERi VE TASARIM KAPASiTESi (Opsiyonel)')

ws_sat.row_dimensions[2].height = 30
ws_sat.merge_cells('A2:E2')
ws_sat['A2'] = ('OPSiYONEL sayfa — doldurulmazsa sistem genel varsayilan degerlerle calisir. '
                'Her AHU icin ozel sogutma/isitma SAT (ufleme) limiti ve tasarim kapasitesi (kW) '
                'girilirse analiz daha dogru olur.')
ws_sat['A2'].fill = PatternFill('solid', start_color=ACIK_MAVI)
ws_sat['A2'].font = Font(size=9, name='Arial', color=KOYU_MAVI, italic=True)
ws_sat['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws_sat.row_dimensions[3].height = 22
for col, text in enumerate(['Mahal', 'AHU Adi', 'Sogutma SAT Limiti (C)',
                              'Isitma SAT Limiti (C)', 'Tasarim Kapasitesi (kW)'], 1):
    c = ws_sat.cell(row=3, column=col)
    c.value = text
    c.fill = PatternFill('solid', start_color=KOYU_MAVI)
    c.font = Font(bold=True, color=BEYAZ, size=10, name='Arial')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for i in range(4, 35):
    ws_sat.row_dimensions[i].height = 20
    for col in range(1, 6):
        ws_sat.cell(row=i, column=col).fill = PatternFill('solid', start_color=SARI)
        ws_sat.cell(row=i, column=col).font = Font(size=10, name='Arial')
        ws_sat.cell(row=i, column=col).alignment = Alignment(horizontal='left', vertical='center')

# Ornek satir
ws_sat.cell(row=4, column=1).value = 'MAS-1'
ws_sat.cell(row=4, column=2).value = 'Ahu-1'
ws_sat.cell(row=4, column=3).value = 18.0
ws_sat.cell(row=4, column=4).value = 28.0
ws_sat.cell(row=4, column=5).value = 194.9
for col in range(1, 6):
    ws_sat.cell(row=4, column=col).font = Font(size=10, name='Arial', italic=True, color='0070C0')

ws_sat.row_dimensions[35].height = 36
ws_sat.merge_cells('A35:E35')
ws_sat['A35'] = ('AHU Adi, "AHU Noktalari" sayfasindaki Point Name icindeki AHU adiyla AYNI olmali '
                  '(orn. "Ahu-1") — eslesmezse o AHU icin ozel limit/kapasite uygulanmaz, genel varsayilan '
                  'kullanilir. Mahal, Lokasyon Profili sayfasindaki Hat 1/Hat 2 adlarina karsilik gelir (MAS-1/MAS-2).')
ws_sat['A35'].fill = PatternFill('solid', start_color=YESIL)
ws_sat['A35'].font = Font(size=9, name='Arial', color='375623', italic=True)
ws_sat['A35'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ════════════════════════════════════════════════
# SAYFA 6 — TALiMATLAR
# ════════════════════════════════════════════════
ws4 = wb.create_sheet('Talimatlar')
ws4.column_dimensions['A'].width = 14
ws4.column_dimensions['B'].width = 75

ws4.row_dimensions[1].height = 30
ws4.merge_cells('A1:B1')
baslik(ws4, 'A1', 'KULLANIM TALiMATLARI', font_size=12)

talimatlar = [
    ('ADIM 1', 'Lokasyon Profili sayfasini doldurun (sari hucreler)'),
    ('ADIM 2', 'Modbus Analizorler sayfasina sahadan aldiginiz IP ve cihaz bilgilerini girin'),
    ('ADIM 3', 'BACnet Noktalar sayfasina Desigo CC santral (chiller/kazan) noktalarini girin'),
    ('ADIM 3b', 'AHU Noktalari sayfasina klima santrali (AHU) noktalarini girin — santral noktalarindan AYRI sayfa'),
    ('ADIM 3c', 'AHU SAT ve Kapasite sayfasini doldurun (OPSiYONEL — daha dogru analiz icin)'),
    ('ADIM 4', 'Doldurulmus Excel dosyasini lokasyon_kurulum_otomasyon.py ile isleyin'),
    ('ADIM 5', 'Otomatik olarak uretilenler (lokasyon_kurulum_otomasyon.py):'),
    ('',       '  - data_collector.py  (Modbus analizor listesiyle, ANALYZERS dizisi)'),
    ('',       '  - data_bridge.py  (Chiller/MCC kategorileri + BACNET_MAP)'),
    ('',       '  - ahu_nokta_konfig.json  (AHU Noktalari sayfasindan)'),
    ('',       '  - ahu_sat_limitleri.json / ahu_tasarim_kapasiteleri.json  (AHU SAT ve Kapasite sayfasindan, doldurulduysa)'),
    ('',       '  - hvac_settings.json  (Genel AHU Ufleme Bandi alanlarindan, CONFIG override — kaynak kod degismez)'),
    ('',       '  - chiller_fcu_ayarlari.json  (Chiller/FCU Kapasite Bilgisi alanlarindan, doldurulduysa)'),
    ('',       '  - location_manager.py  (KRiTiK — lokasyon kimligi + tek/cift hat semasi, Hat Sayisina gore)'),
    ('',       '  - supabase_config.json  (lokasyon ID + Supabase baglanti bilgisi)'),
    ('',       '  - hedefli_okuma_sablonu_2.xlsx  (BACnet Noktalar sayfasindan, Sistem Rolu haric)'),
    ('',       '  - Kurulum zip paketi  (lokasyonlar/<id>.zip)'),
    ('',       ''),
    ('HENUZ YOK', 'Asagidakiler SU AN OTOMATIK DEGIL, manuel yapilmasi gerekiyor:'),
    ('',       '  - Supabase lokasyonlar tablosuna kayit'),
    ('',       '  - app_merkez.py guncellemesi (yeni lokasyon eklenir)'),
    ('',       '  - Hazir kurulum zip paketi'),
    ('',       ''),
    ('ONEMLI', '* isareti zorunlu alandir, digerleri bos birakilabilir'),
    ('ONEMLI', 'IP adresleri sahada kontrol edilmeli — ornek degerleri silin, gercek degerleri girin'),
    ('ONEMLI', 'Modbus veya BACnet yoksa ilgili sayfayi bos birakin'),
    ('ONEMLI', 'Point Name sutunundaki degerler data_collector ile birebir eslesmeli'),
    ('ONEMLI', 'Cihaz Adi BUYUK HARF + tire formatinda olmali (orn. MCC-1, CHILLER-2) — Modbus Analizorler sayfasindaki uyariya bakin'),
    ('ONEMLI', 'Veri Okuma Saati / GM Sync Saati alanlari bilgi amaclidir, sistemde sabittir (07:10 / 08:00)'),
]

for i, (adim, text) in enumerate(talimatlar, 3):
    ws4.row_dimensions[i].height = 22
    if adim.startswith('ADIM'):
        bg_a, fg_a = KOYU_MAVI, BEYAZ
        bg_b = ACIK_MAVI
    elif adim == 'ONEMLI':
        bg_a, fg_a = TURUNCU, '833C00'
        bg_b = TURUNCU
    elif adim == 'HENUZ YOK':
        bg_a, fg_a = GRI, '666666'
        bg_b = GRI
    else:
        bg_a, fg_a = BEYAZ, BEYAZ
        bg_b = BEYAZ

    c1 = ws4.cell(row=i, column=1)
    c1.value = adim if adim else ''
    c1.fill = PatternFill('solid', start_color=bg_a)
    c1.font = Font(bold=True, size=10, name='Arial', color=fg_a)
    c1.alignment = Alignment(horizontal='center', vertical='center')

    c2 = ws4.cell(row=i, column=2)
    c2.value = text
    c2.fill = PatternFill('solid', start_color=bg_b)
    c2.font = Font(size=10, name='Arial')
    c2.alignment = Alignment(horizontal='left', vertical='center')

# Kaydet
dosya = r'C:\Users\furka\OneDrive\Masaüstü\hvac deneme ve gelıstırme program dosyası\yeni_lokasyon_kurulum_sablonu.xlsx'
wb.save(dosya)
print('Olusturuldu:', dosya)
