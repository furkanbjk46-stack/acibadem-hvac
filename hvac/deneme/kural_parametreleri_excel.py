# -*- coding: utf-8 -*-
import os
import pandas as pd

OUT = os.path.join(os.path.dirname(__file__), "static", "outputs", "kural_parametreleri_raporu.xlsx")

df1 = pd.DataFrame([
    ["AHU SOĞUTMA HEDEFİ", "DİNAMİK: emiş−16.5, 4-8°C bandı", "AHU hava ΔT hedefi artık sabit değil; emişe göre hesaplanır (fallback TARGET_AIR_DT_AHU_COOL=6)", "Inlet/Outlet (coil su giriş-çıkış); yoksa Plant Supply/Return. AHU+soğutma+OAT varsa ±2°C OAT bias eklenir"],
    ["TARGET_DT_FCU", "5.0°C", "FCU hedef ΔT (soğutma modu)", "Inlet/Outlet (coil su) -> fallback Plant Supply/Return"],
    ["TARGET_DT_CHILLER", "5.0°C", "Chiller hedef ΔT", "Plant Supply/Return (chillerda Inlet/Outlet genelde yok)"],
    ["TARGET_DT_COLLECTOR", "3.0°C", "Kolektör (ana toplayıcı) hedef ΔT", "Plant Supply (°C) ve Plant Return (°C) - kolektör giriş/çıkış sıcaklıkları. delta_t = |Plant Return - Plant Supply|"],
    ["TARGET_DT_HEAT_EXCHANGER", "8.0°C", "Eşanjör hedef ΔT", "Inlet/Outlet (birincil/ikincil devre su sıcaklıkları)"],
    ["TARGET_DT_HEAT", "15.0°C", "Isıtma devresi (chiller hariç, AHU/FCU dışı tipler) hedef ΔT", "Inlet/Outlet su sıcaklıkları, ısıtma modunda ΔT = Inlet - Outlet"],
    ["(AHU/FCU ısıtma)", "10.0°C", "AHU/FCU coil ısıtma modunda hedef ΔT (özel durum)", "Inlet/Outlet, Mode = Heating"],
    ["TOLERANCE_CRITICAL", "±2.0°C (kodda su an KULLANILMIYOR)", "Parametre analiz fonksiyonlarina tasinir ama hicbir kuralda kullanilmaz; bant kararlari TOLERANCE_NORMAL ile verilir", "Hesaplanan delta_t vs target_delta_t farkı"],
    ["TOLERANCE_NORMAL", "±3.0°C", "Normal tolerans bandı (BAND_LOW/BAND_HIGH/IN_BAND eşiği - FCU)", "Hesaplanan delta_t vs target_delta_t farkı"],
], columns=["Parametre", "Değer", "Açıklama", "Referans Aldığı Veri"])

df2 = pd.DataFrame([
    ["SAT_COOLING_MIN / MAX", "15.0°C / 18.0°C", "SAT (°C) sahası (üfleme sıcaklığı); soğutma modunda bu bandın dışına çıkarsa SAT_LOW/SAT_HIGH"],
    ["SAT_HEATING_MIN / MAX", "28.0°C / 31.0°C", "SAT (°C), ısıtma modunda"],
    ["SAT_COOLING/HEATING_THRESHOLD", "±1.0°C", "SAT (°C) vs Set (°C) karşılaştırması (set'ten sapma toleransı)"],
    ["HEAT_SAT_LOW_THRESHOLD", "28.0°C", "SAT (°C) - LOW_FLOW_DETECTED tetikleyicisi: ΔT (Inlet/Outlet, su) hedefi karşılıyor AMA SAT bu değerin altındaysa -> su ısınıyor ama havaya aktarılamıyor (debi/pompa şüpheli). Çalışan config (hvac_settings.json) değeri"],
], columns=["Parametre", "Değer", "Referans Aldığı Veri"])

df3 = pd.DataFrame([
    ["HIGH_VALVE_THRESHOLD", "%90", "Heat Valve (%) ve Cool Valve (%) - vana açıklık sinyali. >=90 ise 'tam açık' kabul edilir (HEAT_EFF_LOW, COOL_EFF_LOW, LOW_DT_SYNDROME tetikleyicisi)"],
    ["VALVE_SIMUL_THRESHOLD", "%5", "Heat Valve (%) VE Cool Valve (%) aynı anda >=5 ise SIMUL_HEAT_COOL (eşzamanlı ısıtma+soğutma)"],
    ["APPROACH_MAX", "10.0°C", "Cool Valve (%) >=%90 ve approach (Supply hava - Inlet su) > 10°C ise COOL_EFF_LOW"],
    ["COMFORT_DEPARTURE", "3.0°C", "Room (°C) vs Set (°C) -> |Room - Set| > 3.0 ise COMFORT_OVERRIDE"],
    ["LOW_DT_THRESHOLD", "3.0°C", "Hesaplanan delta_t (Inlet/Outlet veya Plant) <=3.0°C iken ilgili vana (ısıtma VEYA soğutma) >=%90 ise LOW_DT_SYNDROME"],
], columns=["Parametre", "Değer", "Referans Aldığı Veri"])

df4 = pd.DataFrame([
    ["CHILLER_BYPASS_DT", "1.0°C", "Plant Supply/Return (veya Inlet/Outlet) farkı < 1.0°C -> su pratikte hiç ısı taşımıyor -> bypass/arıza şüpheli"],
    ["CHILLER_LOW_DT_THRESHOLD", "3.0°C", "Aynı ΔT kaynağı, < 3.0°C -> düşük verimlilik"],
], columns=["Parametre", "Değer", "Referans Aldığı Veri"])

df5 = pd.DataFrame([
    ["OAT_BIAS_MAX", "±2.0°C", "OAT (°C) dış hava sıcaklığı - düşük dış hava sıcaklığında hedef ΔT azaltılır, yüksekte artırılır (AHU hedef ΔT'sine eklenir)"],
], columns=["Parametre", "Değer", "Referans Aldığı Veri"])

df6 = pd.DataFrame([
    ["SCORE_DEPARTURE_WEIGHT", "2.0", "(delta_t - target_delta_t) farkı x 2.0 (kod sabiti, üst sınır +6.0) -> skora eklenir"],
    ["SCORE_LOW_DT_BONUS", "+4.0", "LOW_DT_SYNDROME tetiklendiğinde ek puan"],
    ["SCORE_COMFORT_PENALTY", "+2.0", "COMFORT_OVERRIDE tetiklendiğinde ek puan (Room vs Set sapması)"],
    ["SCORE_CRITICAL_THRESHOLD", "7.0", "Toplam skor >=7.0 -> genel kategori CRITICAL"],
], columns=["Parametre", "Değer", "Referans Aldığı Veri"])

df7 = pd.DataFrame([
    ["FAN_BASMIYOR", "9.5", "CRITICAL", "Sadece AHU (basınç noktalı)", "Start=AÇIK + kanal basıncı <=20 Pa (2 ardışık okuma) -> fan hava basmıyor; analiz atlanır, alarm üretilir"],
    ["TERS_DT", "8.5", "CRITICAL", "Sadece AHU", "Soğutmada hava ΔT < -1.0°C (üfleme emişten sıcak) -> ısıtma kaçağı/sensör karışıklığı"],
    ["LOKAL_CALISMA", "6.0", "WARNING", "Sadece AHU (start+basınç noktalı)", "Start=KAPALI + basınç >20 Pa -> BMS dışı lokal çalıştırma; analiz yine yapılır"],
    ["VERI_EKSIK", "5.0", "WARNING", "AHU", "SAT hedef dışı + EMİŞ verisi yok -> kritik teşhis doğrulanamaz (eski sürüm yanlış KRİTİK basıyordu)"],
    ["SIMUL_HEAT_COOL", "10.0", "CRITICAL", "AHU + FCU", "Heat Valve (%) ve Cool Valve (%) >= %5"],
    ["CHILLER_BYPASS", "9.0", "CRITICAL", "Sadece CHILLER", "Plant/Inlet-Outlet ΔT < 1.0°C"],
    ["NOT_COOLING", "9.0", "CRITICAL", "AHU + FCU", "SAT (°C), Mode=Cooling, SAT > Set+tol, vana >=%70"],
    ["NOT_HEATING", "9.0", "CRITICAL", "AHU + FCU", "SAT (°C), Mode=Heating, SAT < Set-tol, vana >=%70"],
    ["LOW_FLOW_DETECTED", "8.0", "CRITICAL", "AHU + FCU (Chiller hariç)", "Inlet/Outlet ΔT >= 15°C VE SAT < 28°C (isitma modu; AHU haric)"],
    ["INSUFFICIENT_CAPACITY", "6.0", "WARNING", "Sadece FCU", "Room (°C) vs Set (°C) sapması + ilgili Valve (%) (>2°C sapma, vana >%80)"],
    ["HEAT_EFF_LOW", "7.0", "CRITICAL", "Sadece AHU", "Heat Valve >=%90 + üfleme (SAT) < SAT_HEATING_MIN (28°C)"],
    ["COOL_EFF_LOW", "7.0", "CRITICAL", "Sadece AHU", "Cool Valve >=%90 + Approach (Supply-Inlet) > APPROACH_MAX (10°C)"],
    ["CHILLER_LOW_DT", "6.0", "WARNING", "Sadece CHILLER", "Plant/Inlet-Outlet ΔT < 3.0°C"],
    ["AIR_DT_LOW_COOL", "5.0", "WARNING", "Sadece AHU", "Cool Valve >=%90 ama hava ΔT < 3.0°C"],
    ["LOW_DT_SYNDROME", "5.0", "WARNING", "AHU + FCU", "İlgili Valve >=%90 ama ΔT <= 3.0°C"],
    ["SAT_WARNING / HIGH / LOW", "5.0", "WARNING", "AHU + FCU (farklı eşikler)", "SAT (°C) vs SAT_MIN/MAX bandı"],
    ["HIGH_DT", "5.0", "WARNING", "Sadece AHU", "ΔT > target_delta_t + tolerans"],
    ["LOW_DT", "4.0", "WARNING", "Sadece AHU", "ΔT < target_delta_t - tolerans"],
    ["COMFORT_OVERRIDE", "4.0", "WARNING", "AHU + FCU", "|Room - Set| > 3.0°C"],
    ["BAND_LOW / BAND_HIGH", "3.0", "WARNING", "Sadece FCU", "ΔT vs target ± 3.0°C bandı"],
    ["IN_BAND / NORMAL", "0.0", "OPTIMAL", "AHU + FCU", "ΔT bant içinde"],
], columns=["Kural", "Skor", "Kategori", "Geçerli Ekipman", "Kullandığı Sahalar / Tetikleyici"])

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    df1.to_excel(writer, sheet_name="1-Genel DT Hedefleri", index=False)
    df2.to_excel(writer, sheet_name="2-SAT Esikleri", index=False)
    df3.to_excel(writer, sheet_name="3-Vana Approach", index=False)
    df4.to_excel(writer, sheet_name="4-Chiller", index=False)
    df5.to_excel(writer, sheet_name="5-OAT Bias", index=False)
    df6.to_excel(writer, sheet_name="6-Skorlama", index=False)
    df7.to_excel(writer, sheet_name="7-Kural Ozeti", index=False)

    # Sutun genislikleri ve basliklari biraz duzenle
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="002B5E", end_color="002B5E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for ws in writer.sheets.values():
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = min(max(max_len + 2, 15), 70)

print("OK:", OUT)
