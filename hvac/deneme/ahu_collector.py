# ahu_collector.py
# AHU BACnet nokta okuma + main_portal.py HVAC analiz motoru entegrasyonu
# Book2 formatındaki Excel'den üretilen JSON konfig kullanılır.
# Her 30 dakikada bir: BACnet oku → JSON satırları oluştur → /api/recommend_json POST → sonucu kaydet
# cloud_sync._heartbeat_loop tarafından çağrılır (veya bağımsız çalıştırılır).

import socket
import struct
import random
import logging
import os
import json
import urllib.request
import urllib.error
from datetime import datetime

logger = logging.getLogger(__name__)

BASE_DIR          = os.path.dirname(os.path.abspath(__file__))
AHU_CONFIG_FILE   = os.path.join(BASE_DIR, "ahu_nokta_konfig.json")
AHU_RESULTS_FILE  = os.path.join(BASE_DIR, "hvac_ahu_analiz_sonuclari.json")
AHU_TRIGGER_FILE  = os.path.join(BASE_DIR, "hvac_analiz_trigger.txt")
AHU_TASARIM_KAPASITE_FILE = os.path.join(BASE_DIR, "configs", "ahu_tasarim_kapasiteleri.json")
URETIM_TUKETIM_GECMIS_FILE = os.path.join(BASE_DIR, "configs", "uretim_tuketim_gecmis.json")
GECMIS_MAKS_KAYIT = 1008  # 10 dk aralıkla ~1 hafta
ENERGY_CSV_TAZELIK_ESIK_SAAT = 3  # bu süreden eski hedefli_enerji_verileri.csv için uyarı

# data_collector.py ile aynı Excel kaynağı — enerji/kolektör/chiller/kazan noktaları
_ENERGY_EXCEL_CONFIGS = os.path.join(BASE_DIR, "configs", "hedefli_okuma_sablonu_2.xlsx")
_ENERGY_EXCEL_ROOT    = os.path.join(BASE_DIR, "hedefli_okuma_sablonu_2.xlsx")
ENERGY_EXCEL_FILE     = _ENERGY_EXCEL_CONFIGS if os.path.exists(_ENERGY_EXCEL_CONFIGS) else _ENERGY_EXCEL_ROOT

MAIN_PORTAL_URL   = "http://localhost:8005/api/recommend_json"
BACNET_PORT       = 47808
BACNET_TIMEOUT    = 4.0
PROP_PRESENT_VAL  = 85


# ================================================================
# NOKTA TİPİ TESPİTİ
# ================================================================

def _detect_point_type(point_name: str) -> str:
    """Nokta adından canonical key türet (main_portal.py HVACUtils uyumlu)."""
    pn = point_name.lower().strip()
    # Sıra önemli: önce daha spesifik kontroller
    if pn.endswith(" set") or pn == "set":
        return "Set (°C)"
    if "fleme" in pn:                        # üfleme sıcaklığı
        return "SAT (°C)"
    if "so" in pn and "van" in pn:           # soğutma vanası
        return "Cool Valve (%)"
    if "van" in pn:                          # ısıtma vanası (soğutma zaten yakalandı)
        return "Heat Valve (%)"
    return "Return (°C)"                     # emüş / emüs sıcaklığı


def _ahu_adi(point_name: str) -> str:
    """Nokta adından AHU adını çıkar (ilk kelime, örn. 'Ahu-3')."""
    parts = point_name.strip().split()
    return parts[0] if parts else "Unknown"


# ================================================================
# KONFİGÜRASYON YÖNETİMİ
# ================================================================

def load_ahu_config() -> list:
    """AHU nokta konfigürasyonunu JSON dosyasından yükle."""
    if not os.path.exists(AHU_CONFIG_FILE):
        logger.warning("AHU konfig dosyası bulunamadı: %s", AHU_CONFIG_FILE)
        logger.warning("Çalıştır: python ahu_collector.py --setup <Book2.xlsx>")
        return []
    with open(AHU_CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def build_config_from_excel(excel_path: str) -> list:
    """
    Book 2 formatındaki Excel'den AHU konfig JSON üret ve kaydet.
    Kolon yapısı: Gateway IP | Network (DNET) | MAC (DADR) | Device Instance ID
                  Object Type | Object Instance | Point Name | LOCATION (MAHAL)
    Bir kez çalıştır — sonuç ahu_nokta_konfig.json olarak saklanır.
    """
    try:
        import pandas as pd
    except ImportError:
        raise RuntimeError("pandas gerekli: pip install pandas openpyxl")

    df = pd.read_excel(excel_path)
    # Gateway IP ve DNET boşsa bir önceki satırdan doldur
    df["Gateway IP"]     = df["Gateway IP"].ffill()
    df["Network (DNET)"] = df["Network (DNET)"].ffill()
    df["MAC (DADR)"]     = df["MAC (DADR)"].ffill()

    config = []
    for _, row in df.iterrows():
        point_name = str(row.get("Point Name", "")).strip()
        if not point_name or point_name.lower() == "nan":
            continue
        try:
            config.append({
                "ahu_adi":    _ahu_adi(point_name),
                "lokasyon":   str(row.get("LOCATION (MAHAL)", "MAS-1")).strip(),
                "nokta_tipi": _detect_point_type(point_name),
                "gateway_ip": str(row.get("Gateway IP", "")).strip(),
                "dnet":       int(float(row.get("Network (DNET)", 0))),
                "mac_hex":    str(row.get("MAC (DADR)", "0x01")).strip(),
                "obj_type":   int(float(row.get("Object Type", 0))),
                "obj_inst":   int(float(row.get("Object Instance", 0))),
            })
        except Exception as e:
            logger.warning("Satır atlandı (%s): %s", point_name, e)

    with open(AHU_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    logger.info("AHU konfig oluşturuldu: %d nokta → %s", len(config), AHU_CONFIG_FILE)
    return config


def load_energy_config() -> list:
    """
    Kolektör/Chiller/Kazan nokta konfigürasyonunu data_collector.py'nin kullandığı
    aynı Excel dosyasından (hedefli_okuma_sablonu_2.xlsx) doğrudan oku.
    Ayrı bir JSON konfig dosyası tutulmaz — tek kaynak Excel'dir.
    """
    if not os.path.exists(ENERGY_EXCEL_FILE):
        logger.warning("Enerji Excel dosyası bulunamadı: %s", ENERGY_EXCEL_FILE)
        return []
    try:
        import pandas as pd
    except ImportError:
        logger.warning("pandas/openpyxl yüklü değil — enerji noktaları canlı okunamayacak")
        return []

    try:
        df = pd.read_excel(ENERGY_EXCEL_FILE)
        df["Gateway IP"]     = df["Gateway IP"].ffill()
        df["Network (DNET)"] = df["Network (DNET)"].ffill()
        df["MAC (DADR)"]     = df["MAC (DADR)"].ffill()
    except Exception as e:
        logger.warning("Enerji Excel okunamadı: %s", e)
        return []

    config = []
    for _, row in df.iterrows():
        point_name = str(row.get("Point Name", "")).strip()
        if not point_name or point_name.lower() == "nan":
            continue
        try:
            config.append({
                "point_name": point_name.upper(),
                "gateway_ip": str(row.get("Gateway IP", "")).strip(),
                "dnet":       int(float(row.get("Network (DNET)", 0))),
                "mac_hex":    str(row.get("MAC (DADR)", "0x01")).strip(),
                "obj_type":   int(float(row.get("Object Type", 0))),
                "obj_inst":   int(float(row.get("Object Instance", 0))),
            })
        except Exception as e:
            logger.warning("Enerji nokta satırı atlandı (%s): %s", point_name, e)

    return config


# ================================================================
# BACNET OKUMA (data_collector.py ile aynı yaklaşım)
# ================================================================

def _pack_length(packet: bytes) -> bytes:
    return packet[:2] + struct.pack(">H", len(packet)) + packet[4:]


def _build_read_property(obj_type: int, obj_inst: int, prop_id: int = PROP_PRESENT_VAL,
                          invoke_id: int = 1, dnet: int = None, dadr: bytes = None) -> bytes:
    bvlc = bytes([0x81, 0x0a, 0x00, 0x00])
    if dnet is not None and dadr is not None:
        npdu = struct.pack(">BBH B", 0x01, 0x20, dnet, len(dadr)) + dadr + b"\xff"
    else:
        npdu = bytes([0x01, 0x00])
    obj_id = (int(obj_type) << 22) | int(obj_inst)
    apdu  = bytes([0x00, 0x04, invoke_id & 0xFF, 0x0c])
    apdu += bytes([0x0c]) + struct.pack(">I", obj_id)
    apdu += bytes([0x19, prop_id])
    return _pack_length(bvlc + npdu + apdu)


def _parse_real_value(data: bytes):
    """BACnet ComplexACK yanıtından float değer çözümle. Hata varsa None döner."""
    try:
        for i in range(4, min(15, len(data))):
            pdu = data[i] & 0xF0
            if pdu in [0x50, 0x60, 0x70]:   # Error / Reject / Abort
                return None
            if pdu == 0x30:
                break
        prop_idx = data.find(b"\x19\x55")
        if prop_idx != -1:
            start = data.find(b"\x3e", prop_idx)
            end   = data.find(b"\x3f", start)
            payload = data[start+1:end] if (start != -1 and end != -1) else data[prop_idx+2:]
        else:
            start = data.rfind(b"\x3e")
            end   = data.rfind(b"\x3f")
            payload = data[start+1:end] if (start != -1 and end != -1) else data[-10:]
        if not payload:
            return None
        tag       = payload[0]
        tag_class = (tag & 0xF0) >> 4
        tag_len   = tag & 0x0F
        idx = 1
        if tag_len == 5:
            tag_len = payload[1] if len(payload) > 1 else 0
            idx = 2
        if tag_class == 4 and len(payload) >= idx + 4:   # REAL
            return round(struct.unpack(">f", payload[idx:idx+4])[0], 2)
        if tag_class == 1:                                # BOOLEAN
            return 1.0 if tag == 0x11 else 0.0
        if tag_class in [2, 9] and len(payload) >= idx + tag_len:  # UINT/ENUM
            return float(int.from_bytes(payload[idx:idx+tag_len], "big", signed=False))
        if tag_class == 3 and len(payload) >= idx + tag_len:       # SINT
            return float(int.from_bytes(payload[idx:idx+tag_len], "big", signed=True))
    except Exception:
        return None
    return None


def _bacnet_oku(gateway_ip: str, dnet: int, mac_hex: str,
                obj_type: int, obj_inst: int) -> float | None:
    """Tek BACnet noktasını oku. Hata / timeout → None."""
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.bind(("0.0.0.0", 0))
    sock.settimeout(BACNET_TIMEOUT)
    try:
        mac_str = str(mac_hex).replace("0x", "").replace(" ", "").strip()
        if len(mac_str) % 2 != 0:
            mac_str = "0" + mac_str
        mac_bytes = bytes.fromhex(mac_str)
        pkt = _build_read_property(
            obj_type, obj_inst,
            invoke_id=random.randint(1, 255),
            dnet=dnet, dadr=mac_bytes
        )
        sock.sendto(pkt, (gateway_ip, BACNET_PORT))
        data, _ = sock.recvfrom(1024)
        return _parse_real_value(data)
    except Exception:
        return None
    finally:
        sock.close()


# ================================================================
# VERİ OKUMA & FORMATLAMA
# ================================================================

def ahu_verileri_oku(config: list) -> dict:
    """
    Tüm AHU noktalarını BACnet ile oku.
    Dönüş: {(ahu_adi, lokasyon): {"SAT (°C)": val, "Return (°C)": val, ...}}
    """
    sonuclar: dict = {}
    toplam = len(config)
    ok = 0
    for i, nokta in enumerate(config):
        key = (nokta["ahu_adi"], nokta["lokasyon"])
        if key not in sonuclar:
            sonuclar[key] = {}
        deger = _bacnet_oku(
            nokta["gateway_ip"], nokta["dnet"], nokta["mac_hex"],
            nokta["obj_type"], nokta["obj_inst"]
        )
        # Sıcaklık noktalarında BACnet'ten 0.0 gelmesi sensör arızasına işaret eder
        # (0°C oda/üfleme sıcaklığı fiziksel olarak imkânsız). Vana değerleri için 0.0 geçerlidir.
        if deger == 0.0 and nokta["nokta_tipi"].endswith("°C)"):
            deger = None
        sonuclar[key][nokta["nokta_tipi"]] = deger
        if deger is not None:
            ok += 1
        if (i + 1) % 50 == 0:
            logger.info("  BACnet ilerleme: %d/%d (başarılı: %d)", i+1, toplam, ok)

    logger.info("BACnet okuma tamamlandı: %d/%d başarılı", ok, toplam)
    return sonuclar


def energy_verileri_oku(config: list) -> dict:
    """
    Kolektör/Chiller/Kazan noktalarını BACnet ile canlı oku.
    Dönüş: {"POINT_NAME": deger_veya_None}
    """
    sonuclar: dict = {}
    toplam = len(config)
    ok = 0
    for nokta in config:
        deger = _bacnet_oku(
            nokta["gateway_ip"], nokta["dnet"], nokta["mac_hex"],
            nokta["obj_type"], nokta["obj_inst"]
        )
        sonuclar[nokta["point_name"]] = deger
        if deger is not None:
            ok += 1

    logger.info("Enerji noktaları BACnet okuma tamamlandı: %d/%d başarılı", ok, toplam)
    return sonuclar


def _csv_tazelik_kontrol(csv_yolu: str = None, esik_saat: float = ENERGY_CSV_TAZELIK_ESIK_SAAT):
    """hedefli_enerji_verileri.csv'nin son satırının Timestamp'i çok eskiyse uyarı logla."""
    if csv_yolu is None:
        csv_yolu = os.path.join(BASE_DIR, "hedefli_enerji_verileri.csv")
    try:
        import csv as _csv
        son_ts = None
        with open(csv_yolu, "r", encoding="utf-8") as f:
            reader = _csv.DictReader(f)
            for row in reader:
                ts = row.get("Timestamp", "").strip()
                if ts:
                    son_ts = ts
        if not son_ts:
            return
        dt = datetime.strptime(son_ts, "%Y-%m-%d %H:%M:%S")
        fark_saat = (datetime.now() - dt).total_seconds() / 3600
        if fark_saat > esik_saat:
            logger.warning(
                "hedefli_enerji_verileri.csv %.1f saat eski (son okuma: %s) — "
                "fallback olarak kullanılacaksa kolektör/chiller/kazan verileri bayat olabilir",
                fark_saat, son_ts
            )
    except Exception as e:
        logger.warning("hedefli_enerji_verileri.csv tazelik kontrolü yapılamadı: %s", e)


def _enerji_nokta_csv_oku(point_name: str, csv_yolu: str = None) -> float | None:
    """hedefli_enerji_verileri.csv'den tek bir Point_Name değerini okur. Yoksa None."""
    if csv_yolu is None:
        csv_yolu = os.path.join(BASE_DIR, "hedefli_enerji_verileri.csv")
    try:
        import csv as _csv
        with open(csv_yolu, "r", encoding="utf-8") as f:
            reader = _csv.DictReader(f)
            for row in reader:
                pn = str(row.get("Point_Name", "")).strip().upper()
                if pn == point_name.upper():
                    try:
                        return float(row.get("Value", ""))
                    except (ValueError, TypeError):
                        pass
    except Exception as e:
        logger.warning("%s okunamadı: %s", point_name, e)
    return None


def sogutma_outlet_oku(enerji_verileri: dict = None, csv_yolu: str = None) -> float | None:
    """
    MAS-1 ve MAS-2 soğutma sıcaklıklarının ortalamasını döndür.
    Önce canlı BACnet (enerji_verileri), yoksa hedefli_enerji_verileri.csv. Veri yoksa None.
    """
    noktalar = ("MAS-1 SOGUTMA SICAKLIK", "MAS-2 SOGUTMA SICAKLIK")
    degerler = []
    for pn in noktalar:
        val = None
        if enerji_verileri:
            val = enerji_verileri.get(pn)
        if val is None:
            val = _enerji_nokta_csv_oku(pn, csv_yolu)
        if val is not None:
            degerler.append(val)
    if degerler:
        ort = round(sum(degerler) / len(degerler), 2)
        logger.debug("Soğutma Outlet ortalaması: %.2f°C (%d nokta)", ort, len(degerler))
        return ort
    return None


def isitma_outlet_oku(enerji_verileri: dict = None, csv_yolu: str = None) -> float | None:
    """
    MAS-1 ve MAS-2 ısıtma sıcaklıklarının (ısıtma kolektör gidiş) ortalamasını döndür.
    Önce canlı BACnet (enerji_verileri), yoksa hedefli_enerji_verileri.csv. Veri yoksa None.
    """
    noktalar = ("MAS-1 ISITMA SICAKLIK", "MAS-2 ISITMA SICAKLIK")
    degerler = []
    for pn in noktalar:
        val = None
        if enerji_verileri:
            val = enerji_verileri.get(pn)
        if val is None:
            val = _enerji_nokta_csv_oku(pn, csv_yolu)
        if val is not None:
            degerler.append(val)
    if degerler:
        ort = round(sum(degerler) / len(degerler), 2)
        logger.debug("Isıtma Outlet ortalaması: %.2f°C (%d nokta)", ort, len(degerler))
        return ort
    return None


def isitma_donus_oku(enerji_verileri: dict = None, csv_yolu: str = None) -> float | None:
    """
    Ortak kolektör ısıtma dönüş sıcaklığını (KOLLEKTOR ISITMA DONUS SICAKLIK) döndür.
    Önce canlı BACnet (enerji_verileri), yoksa hedefli_enerji_verileri.csv. Veri yoksa None.
    """
    pn = "KOLLEKTOR ISITMA DONUS SICAKLIK"
    if enerji_verileri and enerji_verileri.get(pn) is not None:
        return enerji_verileri[pn]
    return _enerji_nokta_csv_oku(pn, csv_yolu)


def kollektor_donus_oku(enerji_verileri: dict = None, csv_yolu: str = None) -> float | None:
    """
    Ortak kolektör soğutma dönüş sıcaklığını (KOLLEKTOR DONUS SICAKLIK) döndür.
    Önce canlı BACnet (enerji_verileri), yoksa hedefli_enerji_verileri.csv. Veri yoksa None.
    """
    adlar = ("KOLLEKTOR SOGUTMA DONUS SICAKLIK", "KOLLEKTOR DONUS SICAKLIK")
    if enerji_verileri:
        for pn in adlar:
            if enerji_verileri.get(pn) is not None:
                return enerji_verileri[pn]
    for pn in adlar:
        val = _enerji_nokta_csv_oku(pn, csv_yolu)
        if val is not None:
            return val
    return None


def _nokta_oku(point_name: str, enerji_verileri: dict = None, csv_yolu: str = None) -> float | None:
    """
    Tek bir nokta değerini döndür (CH-x / KAZAN-x DURUM/INLET/OUTLET vb.).
    Önce canlı BACnet (enerji_verileri), yoksa hedefli_enerji_verileri.csv. Yoksa None.
    """
    if enerji_verileri:
        val = enerji_verileri.get(point_name.upper())
        if val is not None:
            return val
    return _enerji_nokta_csv_oku(point_name, csv_yolu)


def ek_ekipman_satirlari_olustur(outlet: float | None, inlet: float | None,
                                  outlet_heat: float | None, inlet_heat: float | None,
                                  lokasyon: str = "MAS-1", enerji_verileri: dict = None) -> list:
    """
    AHU dışındaki merkezi ekipmanlar için analiz satırları:
    - Chiller        : soğutma kolektör gidiş/dönüş (TARGET_DT_CHILLER ~5°C)
    - Isıtma Kolektörü: ısıtma kolektör gidiş/dönüş (TARGET_DT_COLLECTOR ~10°C)
    - Kazan-1/2/3    : kendi INLET/OUTLET noktaları (TARGET_DT_HEAT_EXCHANGER ~8°C)
    """
    rows = []

    CALISMA_DT_ESIGI = 0.5  # °C — bu değerin altındaki ΔT, cihaz çalışmıyor demektir

    # CH-1..5 chiller'ları — kendi DURUM BILGISI ve INLET/OUTLET noktalarıyla
    CH_OUTLET_ADLARI = {
        1: "CH-1 OUTLET CIKIS SICAKLIK",
        2: "CH-2 OUTLET SICAKLIK",
        3: "CH-3 OUTLET SICAKLIK",
        4: "CH-4 OUTLET SICAKLIK",
        5: "CH-5 OUTLET CIKIS SICAKLIK",
    }
    for n in (1, 2, 3, 4, 5):
        durum = _nokta_oku(f"CH-{n} DURUM BILGISI", enerji_verileri)
        if durum is not None and durum == 0:
            logger.info("CH-%d kapalı (DURUM=0) — analiz atlandı", n)
            continue

        c_inlet = _nokta_oku(f"CH-{n} INLET SICAKLIK", enerji_verileri)
        c_outlet = _nokta_oku(CH_OUTLET_ADLARI[n], enerji_verileri)
        if c_inlet is not None and c_outlet is not None:
            if abs(c_outlet - c_inlet) < CALISMA_DT_ESIGI:
                logger.info("CH-%d ΔT≈0 — çalışmıyor, analiz atlandı", n)
                continue
            row = {
                "Name": f"CH-{n}", "Location": lokasyon, "Type": "Chiller",
                "Inlet (°C)": c_inlet, "Outlet (°C)": c_outlet,
            }
            yuzde = _nokta_oku(f"CH-{n} CALISMA YUZDELIK", enerji_verileri)
            if yuzde is not None:
                row["Calisma Yuzdelik (%)"] = yuzde
            ic_set = _nokta_oku(f"CH-{n} IC SET", enerji_verileri)
            if ic_set is not None:
                row["Set (°C)"] = ic_set
            rows.append(row)

    if outlet_heat is not None and inlet_heat is not None and abs(outlet_heat - inlet_heat) >= CALISMA_DT_ESIGI:
        rows.append({
            "Name": "Isitma Kollektoru", "Location": lokasyon, "Type": "Collector",
            "Inlet (°C)": inlet_heat, "Outlet (°C)": outlet_heat,
        })

    for n in (1, 2, 3):
        durum = _nokta_oku(f"KAZAN-{n} DURUM BILGISI", enerji_verileri)
        if durum is not None and durum == 0:
            logger.info("Kazan-%d kapalı (DURUM=0) — analiz atlandı", n)
            continue

        k_inlet = _nokta_oku(f"KAZAN-{n} INLET", enerji_verileri)
        k_outlet = _nokta_oku(f"KAZAN-{n} OUTLET", enerji_verileri)
        if k_inlet is not None and k_outlet is not None:
            if abs(k_outlet - k_inlet) < CALISMA_DT_ESIGI:
                logger.info("Kazan-%d ΔT≈0 — çalışmıyor, analiz atlandı", n)
                continue
            rows.append({
                "Name": f"Kazan-{n}", "Location": lokasyon, "Type": "Heat Exchanger",
                "Inlet (°C)": k_inlet, "Outlet (°C)": k_outlet,
            })

    return rows


def chiller_uretim_hesapla(enerji_verileri: dict = None) -> dict:
    """
    Çalışan chiller'ların yük yüzdesine (CH-n CALISMA YUZDELIK) ve Sistem
    Ayarları'ndaki birim kapasiteye (chiller_birim_kw) göre gerçek üretilen
    soğutma kW'ını hesaplar. Debi ölçümü olmadığı için bu, su ΔT'sinden değil
    chiller'ın kendi bildirdiği yük yüzdesinden türetilir.

    fcu_kapasite_kw, Sistem Ayarları'ndaki FCU filosunun (fcu_adedi x
    fcu_birim_kw_ortalama x fcu_esanjor_diversity) tüketim/dağıtım tarafındaki
    teorik kapasitesidir — bilgi amaçlıdır, üretim değerini sınırlamaz
    (üretim tamamen chiller tarafına ait bir değer).
    Dönen: {"uretilen_kw": float, "calisan_adet": int, "ortalama_yuzde": float|None,
            "fcu_kapasite_kw": float}
    """
    try:
        chiller_fcu_dosya = os.path.join(BASE_DIR, "configs", "chiller_fcu_ayarlari.json")
        with open(chiller_fcu_dosya, "r", encoding="utf-8") as f:
            ayarlar = json.load(f)
        birim_kw = float(ayarlar.get("chiller_birim_kw", 0))
        chiller_adedi = int(ayarlar.get("chiller_adedi", 5))
        fcu_adedi = float(ayarlar.get("fcu_adedi", 0))
        fcu_birim_kw = float(ayarlar.get("fcu_birim_kw_ortalama", 0))
        fcu_diversity = float(ayarlar.get("fcu_esanjor_diversity", 1.0))
    except Exception:
        birim_kw = 0
        chiller_adedi = 5
        fcu_adedi = 0
        fcu_birim_kw = 0
        fcu_diversity = 1.0

    fcu_kapasite_kw = fcu_adedi * fcu_birim_kw * fcu_diversity

    uretilen_kw = 0.0
    calisan_adet = 0
    yuzdeler = []

    # BACnet şablonunda CH-1..CH-5 tanımlı — lokasyondaki gerçek chiller adedi
    # (Sistem Ayarları'ndan) bu üst sınırı aşamaz
    for n in range(1, min(chiller_adedi, 5) + 1):
        durum = _nokta_oku(f"CH-{n} DURUM BILGISI", enerji_verileri)
        if durum is not None and durum == 0:
            continue
        yuzde = _nokta_oku(f"CH-{n} CALISMA YUZDELIK", enerji_verileri)
        if yuzde is None:
            continue
        calisan_adet += 1
        yuzdeler.append(yuzde)
        uretilen_kw += birim_kw * (yuzde / 100.0)

    ortalama_yuzde = sum(yuzdeler) / len(yuzdeler) if yuzdeler else None
    return {
        "uretilen_kw": round(uretilen_kw, 1),
        "calisan_adet": calisan_adet,
        "ortalama_yuzde": round(ortalama_yuzde, 1) if ortalama_yuzde is not None else None,
        "fcu_kapasite_kw": round(fcu_kapasite_kw, 1),
    }


def talep_hesapla(rows: list) -> float:
    """
    AHU'ların soğutma vana açıklığı (%) ve Sistem Ayarları'ndaki tasarım
    kapasitelerine (ahu_tasarim_kapasiteleri.json) göre talep edilen soğutma
    kW'ını hesaplar (frontend'deki updateCoolingKw ile aynı mantık).
    FCU filosu, chiller -> kolektör -> FCU hiyerarşisinde üretilen soğutmayı
    tüketen taraf olduğu için talebe ayrıca eklenmez/tavan olarak kullanılmaz
    (çift sayım olur) — fcu_kapasite_kw sadece bilgi amaçlı ayrı tutulur.
    """
    try:
        with open(AHU_TASARIM_KAPASITE_FILE, "r", encoding="utf-8") as f:
            kapasiteler = json.load(f)
    except Exception:
        kapasiteler = {}

    talep_kw = 0.0
    for r in rows:
        if (r.get("Type") or "AHU").upper() != "AHU":
            continue
        cv = r.get("Cool Valve (%)")
        if cv is None or cv <= 0:
            continue
        qt = (kapasiteler.get(r.get("Location", "")) or {}).get(r.get("Name", ""))
        if qt is None:
            continue
        talep_kw += qt * (cv / 100.0)

    return round(talep_kw, 1)


def gecmis_kaydet(uretilen_kw: float, talep_kw: float) -> None:
    """Üretilen/talep edilen kW'ı zaman damgasıyla geçmiş dosyasına ekler (canlı grafik için)."""
    try:
        os.makedirs(os.path.dirname(URETIM_TUKETIM_GECMIS_FILE), exist_ok=True)
        try:
            with open(URETIM_TUKETIM_GECMIS_FILE, "r", encoding="utf-8") as f:
                gecmis = json.load(f)
        except Exception:
            gecmis = []

        gecmis.append({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "uretilen_kw": uretilen_kw,
            "talep_kw": talep_kw,
        })
        gecmis = gecmis[-GECMIS_MAKS_KAYIT:]

        with open(URETIM_TUKETIM_GECMIS_FILE, "w", encoding="utf-8") as f:
            json.dump(gecmis, f, ensure_ascii=False)
    except Exception as e:
        logger.error("Üretim/tüketim geçmişi kaydedilemedi: %s", e)


def _ahu_row_olustur(ahu_adi: str, lokasyon: str, veriler: dict,
                     outlet: float | None = None, inlet: float | None = None,
                     outlet_heat: float | None = None, inlet_heat: float | None = None) -> dict:
    """
    AHU verisini /api/recommend_json formatına dönüştür.
    - Room    = emiş sıcaklığı (Return ile aynı)
    - Outlet/Inlet (soğutma) = MAS-1/MAS-2 soğutma gidiş ortalaması / kolektör soğutma dönüş
    - Outlet/Inlet (ısıtma)  = MAS-1/MAS-2 ısıtma gidiş ortalaması / kolektör ısıtma dönüş

    Hangi çiftin kullanılacağı AHU'nun Heat Valve / Cool Valve açıklığına göre
    seçilir: Heat Valve > Cool Valve ise ısıtma kolektör çifti, aksi halde
    soğutma kolektör çifti atanır.
    """
    row: dict = {
        "Name":     ahu_adi,
        "Location": lokasyon,
        "Type":     "AHU",
    }

    # BACnet'ten okunan değerler
    for alan in ["SAT (°C)", "Return (°C)", "Cool Valve (%)", "Heat Valve (%)", "Set (°C)"]:
        if veriler.get(alan) is not None:
            row[alan] = veriler[alan]

    # Room = emiş (Return) ile aynı
    if veriler.get("Return (°C)") is not None:
        row["Room (°C)"] = veriler["Return (°C)"]

    # AHU modu: Heat Valve > Cool Valve ise ısıtma kolektör çiftini kullan
    cool_valve = veriler.get("Cool Valve (%)") or 0
    heat_valve = veriler.get("Heat Valve (%)") or 0
    if heat_valve > cool_valve and inlet_heat is not None and outlet_heat is not None:
        row["Inlet (°C)"] = inlet_heat
        row["Outlet (°C)"] = outlet_heat
    else:
        if inlet is not None:
            row["Inlet (°C)"] = inlet
        if outlet is not None:
            row["Outlet (°C)"] = outlet

    return row


# ================================================================
# OAT (DIŞ HAVA) — Open-Meteo'dan anlık sıcaklık
# ================================================================

def _oat_cek(lat: float = 41.1, lon: float = 29.0) -> float:
    """Open-Meteo'dan anlık dış hava sıcaklığını çek. Hata → 20.0."""
    try:
        url = (
            f"https://api.open-meteo.com/v1/forecast"
            f"?latitude={lat}&longitude={lon}"
            f"&current=temperature_2m&forecast_days=1"
        )
        with urllib.request.urlopen(url, timeout=8) as r:
            data = json.loads(r.read().decode())
        return float(data["current"]["temperature_2m"])
    except Exception as e:
        logger.warning("OAT çekilemedi: %s — varsayılan 20°C", e)
        return 20.0


# ================================================================
# ANALİZ API ENTEGRASYONU
# ================================================================

def analiz_gonder(rows: list, oat: float = 20.0, chiller_load_percent: float | None = None) -> dict | None:
    """
    main_portal.py /api/recommend_json endpoint'ine POST et.
    Başarısız → None.
    """
    body = {
        "rows":   rows,
        "oat":    str(oat),
        "engine": "v2",
    }
    if chiller_load_percent is not None:
        body["chiller_load_percent"] = str(chiller_load_percent)
    payload = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        MAIN_PORTAL_URL,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        logger.error("HVAC analiz API hatası: %s", e)
        return None


def bacnet_okuma_csv_kaydet(veriler: dict, zaman_damgasi: str) -> str:
    """
    Ham BACnet okuma sonuçlarını CSV'ye yaz.
    Format: Timestamp, AHU_Adi, Lokasyon, Nokta_Tipi, Deger, Durum
    Dönüş: Kaydedilen dosya adı
    """
    import csv as _csv
    klasor = os.path.join(BASE_DIR, "ahu_okuma_gecmis")
    os.makedirs(klasor, exist_ok=True)
    dosya_adi = f"ahu_bacnet_{zaman_damgasi}.csv"
    tam_yol   = os.path.join(klasor, dosya_adi)

    with open(tam_yol, "w", newline="", encoding="utf-8") as f:
        yazar = _csv.writer(f)
        yazar.writerow(["Timestamp", "AHU_Adi", "Lokasyon", "Nokta_Tipi", "Deger", "Durum"])
        for (ahu_adi, lokasyon), noktalar in sorted(veriler.items()):
            for nokta_tipi, deger in noktalar.items():
                durum = "OK" if deger is not None else "Zaman Asimi"
                yazar.writerow([
                    zaman_damgasi.replace("_", " "),
                    ahu_adi, lokasyon, nokta_tipi,
                    round(deger, 2) if deger is not None else "",
                    durum
                ])

    logger.info("BACnet okuma CSV kaydedildi → %s (%d AHU)", tam_yol, len(veriler))
    return dosya_adi


def analiz_detay_csv_kaydet(results: list, zaman_damgasi: str) -> str:
    """
    Analiz sonuçlarını personel için detaylı CSV'ye yaz (web arayüzündeki tablo ile aynı).
    Dönüş: Kaydedilen dosya adı
    """
    import csv as _csv
    klasor = os.path.join(BASE_DIR, "ahu_analiz_gecmis")
    os.makedirs(klasor, exist_ok=True)
    dosya_adi = f"hvac_analysis_{zaman_damgasi}.csv"
    tam_yol   = os.path.join(klasor, dosya_adi)

    # Web arayüzündeki tabloyla birebir aynı sütunlar
    sutunlar = [
        "Mahal", "Tip", "Ekipman", "Aksiyon",
        "Delta_T_C", "Hedef_C", "Sapma",
        "SAT_Durum", "Onerilen_SAT_C",
        "Skor", "Durum", "Siddет",
        "Yaklasim_Giris", "Yaklasim_Donus",
        "Kural", "Talimat_Baslik",
        "Ufleme_SAT_C", "Emis_C", "Set_C",
        "Sogutma_Vana_Yuzde", "Isitma_Vana_Yuzde",
    ]

    with open(tam_yol, "w", newline="", encoding="utf-8") as f:
        yazar = _csv.DictWriter(f, fieldnames=sutunlar, extrasaction="ignore")
        yazar.writeheader()
        for r in results:
            yazar.writerow({
                "Mahal":              r.get("Location", ""),
                "Tip":                r.get("Type", ""),
                "Ekipman":            r.get("Name", r.get("Asset", "")),
                "Aksiyon":            r.get("Action", ""),
                "Delta_T_C":          r.get("ΔT (°C)", r.get("Su ΔT (°C)", "")),
                "Hedef_C":            r.get("Target ΔT (°C)", r.get("Hedef", "")),
                "Sapma":              r.get("Departure", r.get("Sapma", "")),
                "SAT_Durum":          r.get("SAT Status", r.get("SAT Durum", "")),
                "Onerilen_SAT_C":     r.get("Recommended SAT (°C)", ""),
                "Skor":               r.get("Score", ""),
                "Durum":              r.get("Status", ""),
                "Siddет":             r.get("Severity", ""),
                "Yaklasim_Giris":     r.get("Approach Gidis", r.get("Approach Supply", "")),
                "Yaklasim_Donus":     r.get("Approach Donus", r.get("Approach Return", "")),
                "Kural":              r.get("Rule", ""),
                "Talimat_Baslik":     r.get("Instruction Title", r.get("instruction_title", "")),
                "Ufleme_SAT_C":       r.get("SAT (°C)", r.get("Supply (°C)", "")),
                "Emis_C":             r.get("Return (°C)", ""),
                "Set_C":              r.get("Set (°C)", ""),
                "Sogutma_Vana_Yuzde": r.get("Cool Valve (%)", ""),
                "Isitma_Vana_Yuzde":  r.get("Heat Valve (%)", ""),
            })

    logger.info("Detaylı analiz CSV kaydedildi → %s (%d satır)", tam_yol, len(results))
    return dosya_adi


def _bildirim_gonder(kritik: int, uyari: int, optimal: int, toplam: int):
    """
    Analiz tamamlanınca Supabase bildirimler tablosuna bildirim ekle.
    Personel app_portal.py'deki banner'da görür.
    """
    try:
        config_yolu = os.path.join(BASE_DIR, "supabase_config.json")
        if not os.path.exists(config_yolu):
            return
        with open(config_yolu, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        sb_url = cfg.get("supabase_url", "")
        sb_key = cfg.get("supabase_key", "")
        lok_id = cfg.get("lokasyon_id", "maslak")
        if not sb_url or not sb_key:
            return

        # Öncelik belirle
        if kritik > 0:
            oncelik = "acil"
        elif uyari > 0:
            oncelik = "uyari"
        else:
            oncelik = "bilgi"

        mesaj = (
            f"🌬️ HVAC Santral Analizi Tamamlandı — "
            f"{toplam} santral | 🔴 {kritik} Kritik | 🟡 {uyari} Uyarı | 🟢 {optimal} Optimal\n"
            f"Sonuçları HVAC portalde kontrol edin ve onaylayın."
        )

        payload = json.dumps({
            "lokasyon": lok_id,
            "mesaj":    mesaj,
            "gonderen": "HVAC Sistem",
            "oncelik":  oncelik,
            "okundu":   False,
        }).encode("utf-8")

        req = urllib.request.Request(
            sb_url + "/rest/v1/bildirimler",
            data=payload,
            headers={
                "apikey":        sb_key,
                "Authorization": "Bearer " + sb_key,
                "Content-Type":  "application/json",
                "Prefer":        "return=minimal",
            },
            method="POST",
        )
        urllib.request.urlopen(req, timeout=8)
        logger.info("Supabase bildirim gönderildi: %s (%s)", oncelik, lok_id)
    except Exception as e:
        logger.warning("Bildirim gönderilemedi: %s", e)


def sonuclari_kaydet(sonuc: dict, veriler: dict, ahu_sayisi: int, oat: float):
    """
    3 çıktı üret + bildirim gönder:
    1) hvac_ahu_analiz_sonuclari.json  — son analizin tam JSON'u (main_portal için)
    2) ahu_okuma_gecmis/ahu_bacnet_YYYYMMDD_HHMM.csv — ham BACnet okuma
    3) ahu_analiz_gecmis/hvac_analysis_YYYYMMDD_HHMM.csv — detaylı analiz tablosu
    NOT: hvac_analysis_history.csv → main_portal.py zaten analiz sırasında kendisi yazıyor.
    """
    zaman_damgasi = datetime.now().strftime("%Y%m%d_%H%M")
    results = sonuc.get("recs_table", {}).get("rows", [])
    kpi     = sonuc.get("kpi", {})

    # 1) JSON
    out = {
        "timestamp":  datetime.now().isoformat(),
        "ahu_sayisi": ahu_sayisi,
        "oat":        oat,
        "sonuc":      sonuc,
    }
    with open(AHU_RESULTS_FILE, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # 2) Ham BACnet okuma CSV
    bacnet_okuma_csv_kaydet(veriler, zaman_damgasi)

    # 3) Detaylı analiz CSV
    detay_dosya_adi = analiz_detay_csv_kaydet(results, zaman_damgasi)

    # 3b) Aylık alarm sayacını güncelle (gün-gün karşılaştırma)
    try:
        from monthly_report.ahu_alarm_takip import gunluk_analiz_isle
        detay_yolu = os.path.join(BASE_DIR, "ahu_analiz_gecmis", detay_dosya_adi)
        karsilastirma = gunluk_analiz_isle(detay_yolu)
        logger.info(
            "Alarm sayacı güncellendi — yeni: %s, devam eden: %s, düzelen: %s",
            karsilastirma["yeni"], karsilastirma["devam_eden"], karsilastirma["duzelen"]
        )
    except Exception as e:
        logger.warning("Alarm sayacı güncellenemedi: %s", e)

    # 4) Supabase bildirim
    _bildirim_gonder(
        kritik  = kpi.get("crit", 0),
        uyari   = kpi.get("warn", 0),
        optimal = kpi.get("optimal", 0),
        toplam  = kpi.get("total", ahu_sayisi),
    )

    logger.info("Tüm çıktılar kaydedildi — %d AHU, zaman: %s", ahu_sayisi, zaman_damgasi)


# ================================================================
# ANA GİRİŞ NOKTASI
# ================================================================

def hvac_analiz_calistir() -> dict | None:
    """
    Tam döngü:
    1) Konfig yükle (ahu_nokta_konfig.json)
    2) BACnet ile tüm AHU noktalarını oku
    3) Satırları /api/recommend_json formatına çevir
    4) main_portal.py'ye POST et
    5) Sonucu hvac_ahu_analiz_sonuclari.json'a kaydet
    6) Trigger dosyası varsa sil
    """
    # Trigger dosyasını temizle (manuel "Yeniden Analiz" isteği)
    try:
        if os.path.exists(AHU_TRIGGER_FILE):
            os.remove(AHU_TRIGGER_FILE)
    except Exception:
        pass

    config = load_ahu_config()
    if not config:
        return None

    logger.info("═══ HVAC AHU ANALİZİ BAŞLADI (%d nokta) ═══", len(config))

    # Dış hava sıcaklığını çek
    oat = _oat_cek()
    logger.info("Dış hava sıcaklığı: %.1f°C", oat)

    # hedefli_enerji_verileri.csv bayatsa uyar (fallback olarak kullanılacaksa)
    _csv_tazelik_kontrol()

    # Kolektör/Chiller/Kazan noktalarını canlı BACnet ile oku (config varsa)
    enerji_config = load_energy_config()
    enerji_verileri = energy_verileri_oku(enerji_config) if enerji_config else None

    # Soğutma suyu Outlet (gidiş) sıcaklığını çek (MAS-1 + MAS-2 ortalaması)
    outlet = sogutma_outlet_oku(enerji_verileri)
    if outlet is not None:
        logger.info("Soğutma Outlet/Gidiş (ortalama): %.2f°C", outlet)
    else:
        logger.warning("Soğutma Outlet/Gidiş okunamadı — Outlet sütunu boş bırakılacak")

    # Kolektör soğutma Inlet (dönüş) sıcaklığını çek (ortak kolektör)
    inlet = kollektor_donus_oku(enerji_verileri)
    if inlet is not None:
        logger.info("Soğutma Inlet/Dönüş (kolektör): %.2f°C", inlet)
    else:
        logger.warning("Soğutma Inlet/Dönüş okunamadı — Inlet sütunu boş bırakılacak")

    # Isıtma suyu Outlet (gidiş) sıcaklığını çek (MAS-1 + MAS-2 ortalaması)
    outlet_heat = isitma_outlet_oku(enerji_verileri)
    if outlet_heat is not None:
        logger.info("Isıtma Outlet/Gidiş (ortalama): %.2f°C", outlet_heat)
    else:
        logger.warning("Isıtma Outlet/Gidiş okunamadı")

    # Kolektör ısıtma Inlet (dönüş) sıcaklığını çek (ortak kolektör)
    inlet_heat = isitma_donus_oku(enerji_verileri)
    if inlet_heat is not None:
        logger.info("Isıtma Inlet/Dönüş (kolektör): %.2f°C", inlet_heat)
    else:
        logger.warning("Isıtma Inlet/Dönüş okunamadı")

    # BACnet okuma
    veriler = ahu_verileri_oku(config)

    # Ham BACnet verilerini debug Excel'e yaz
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font

        # Tüm nokta tiplerini topla (sütun başlıkları)
        tum_nokta_tipleri = sorted({
            pt for nokta_verileri in veriler.values()
            for pt in nokta_verileri.keys()
        })

        wb_debug = openpyxl.Workbook()
        ws_debug = wb_debug.active
        ws_debug.title = "Ham BACnet Verileri"

        basliklar = ["AHU", "Lokasyon"] + tum_nokta_tipleri
        ws_debug.append(basliklar)

        # Başlık satırı formatla
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws_debug[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)

        # None = kırmızı, değer var = beyaz
        none_fill = PatternFill("solid", fgColor="FFCCCC")

        for (ahu_adi, lokasyon), nokta_verileri in sorted(veriler.items()):
            satir = [ahu_adi, lokasyon] + [
                nokta_verileri.get(pt) for pt in tum_nokta_tipleri
            ]
            ws_debug.append(satir)
            row_idx = ws_debug.max_row
            for col_idx, pt in enumerate(tum_nokta_tipleri, start=3):
                if nokta_verileri.get(pt) is None:
                    ws_debug.cell(row=row_idx, column=col_idx).fill = none_fill

        # Enerji verileri — ayrı sheet
        if enerji_verileri:
            ws_enerji = wb_debug.create_sheet("Enerji Ham Verileri")
            ws_enerji.append(["Nokta Adı", "Değer", "Durum"])
            for cell in ws_enerji[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
            for nokta_adi, deger in sorted(enerji_verileri.items()):
                durum = "OK" if deger is not None else "OKUNAMADI"
                ws_enerji.append([nokta_adi, deger, durum])
                if deger is None:
                    for col in range(1, 4):
                        ws_enerji.cell(row=ws_enerji.max_row, column=col).fill = none_fill

        debug_dosya = os.path.join(BASE_DIR, "bacnet_ham_veriler.xlsx")
        wb_debug.save(debug_dosya)
        logger.info("Ham BACnet verileri kaydedildi: %s", debug_dosya)
    except Exception as _e:
        logger.warning("Ham BACnet debug Excel yazılamadı: %s", _e)

    # Satırları oluştur (SAT yoksa AHU analiz edilemez)
    rows = []
    eksik = 0
    for (ahu_adi, lokasyon), nokta_verileri in veriler.items():
        if nokta_verileri.get("SAT (°C)") is None:
            eksik += 1
            continue
        rows.append(_ahu_row_olustur(ahu_adi, lokasyon, nokta_verileri, outlet=outlet, inlet=inlet,
                                      outlet_heat=outlet_heat, inlet_heat=inlet_heat))

    # Merkezi ekipmanlar: Chiller, Isıtma Kolektörü, Kazan-1/2/3
    rows.extend(ek_ekipman_satirlari_olustur(outlet, inlet, outlet_heat, inlet_heat,
                                              enerji_verileri=enerji_verileri))

    logger.info(
        "AHU: toplam=%d, eksik SAT=%d, analiz edilecek=%d",
        len(veriler), eksik, len(rows)
    )

    if not rows:
        logger.warning("Analiz edilecek AHU yok — BACnet bağlantısını kontrol edin.")
        return None

    # Chiller yük yüzdesinden gerçek üretilen soğutma kW'ını hesapla
    uretim = chiller_uretim_hesapla(enerji_verileri)
    logger.info(
        "Chiller üretimi: %.0f kW (%d çalışan, ortalama yük %s%%)",
        uretim["uretilen_kw"], uretim["calisan_adet"], uretim["ortalama_yuzde"]
    )

    # API'ye gönder
    sonuc = analiz_gonder(rows, oat=oat, chiller_load_percent=uretim["ortalama_yuzde"])
    if sonuc:
        talep_kw = talep_hesapla(rows)
        sonuc["uretilen_sogutma_kw"] = uretim["uretilen_kw"]
        sonuc["calisan_chiller_adet"] = uretim["calisan_adet"]
        sonuc["chiller_ortalama_yuk_pct"] = uretim["ortalama_yuzde"]
        sonuc["fcu_kapasite_kw"] = uretim["fcu_kapasite_kw"]
        sonuc["talep_sogutma_kw"] = talep_kw
        sonuclari_kaydet(sonuc, veriler=veriler, ahu_sayisi=len(rows), oat=oat)
        gecmis_kaydet(uretim["uretilen_kw"], talep_kw)
        logger.info("═══ HVAC ANALİZİ TAMAMLANDI (%d AHU) ═══", len(rows))
    return sonuc


# ================================================================
# KOMMAnd-LINE: python ahu_collector.py --setup <excel_dosyasi>
# ================================================================

if __name__ == "__main__":
    import sys

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [AHU] %(message)s",
        datefmt="%H:%M:%S",
    )

    if len(sys.argv) >= 3 and sys.argv[1] == "--setup":
        excel_yolu = sys.argv[2]
        print(f"Excel okunuyor: {excel_yolu}")
        cfg = build_config_from_excel(excel_yolu)
        print(f"Konfig oluşturuldu: {len(cfg)} nokta → {AHU_CONFIG_FILE}")
    else:
        print("Kullanım:")
        print("  AHU konfig oluştur : python ahu_collector.py --setup <Book2.xlsx>")
        print("  Analiz çalıştır: import ahu_collector; ahu_collector.hvac_analiz_calistir()")
        print("  (Enerji/kolektör noktaları otomatik olarak hedefli_okuma_sablonu_2.xlsx'ten okunur)")
        print()
        print("Tek seferlik analiz başlatılıyor...")
        hvac_analiz_calistir()
