# lokasyon_kurulum_otomasyon.py
# Doldurulmus "yeni_lokasyon_kurulum_sablonu.xlsx" dosyasini okur, yeni lokasyon icin
# baslangic dosyalarini uretir.
#
# KAPSAM (su an otomatik olan):
#   - data_collector.py  (Modbus Analizorler sayfasindan ANALYZERS listesi uretilir)
#   - data_bridge.py     (CHILLER_ANALYZERS/ALL_ANALYZERS + BACNET_MAP, Modbus + BACnet
#                          sayfalarindan, Sistem Rolu eslemesiyle uretilir)
#   - supabase_config.json  (Lokasyon Profili sayfasindan lokasyon_id uretilir)
#   - ahu_nokta_konfig.json  (AHU Noktalari sayfasindan, ahu_collector.py --setup ile ayni format)
#   - ahu_sat_limitleri.json / ahu_tasarim_kapasiteleri.json  (AHU SAT ve Kapasite sayfasindan, opsiyonel)
#   - configs/hvac_settings.json  (Genel AHU Ufleme Bandi alanlarindan — main_portal.py
#                                   acilista bunu okuyup CONFIG'i override eder, kaynak kod degismez)
#   - configs/chiller_fcu_ayarlari.json  (Chiller/FCU Kapasite Bilgisi alanlarindan, doldurulduysa)
#   - location_manager.py  (KRiTiK — app_portal.py'nin lokasyon kimligi + tek/cift hat
#                            semasi buradan gelir; Hat Sayisi alanina gore otomatik uretilir)
#   - configs/hedefli_okuma_sablonu_2.xlsx  (BACnet Noktalar sayfasindan, Sistem Rolu haric)
#   - Kurulum zip paketi  (lokasyonlar/<id>.zip)
#   - Supabase 'lisanslar' tablosuna kayit  (Makine ID alani doldurulduysa)
#
# KAPSAM DISI (henuz otomatik degil, manuel yapilmali):
#   - Supabase 'lokasyonlar' tablosuna kayit
#   - app_merkez.py guncellemesi (HASTANELER dict'ine yeni lokasyon eklenir)
#
# Kullanim:
#   python lokasyon_kurulum_otomasyon.py yeni_lokasyon_kurulum_sablonu.xlsx

import sys
import os
import re
import json
import shutil
from openpyxl import load_workbook, Workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
COLLECTOR_SABLON = os.path.join(BASE_DIR, "hvac", "deneme", "data_collector.py")
BRIDGE_SABLON = os.path.join(BASE_DIR, "hvac", "deneme", "data_bridge.py")

GECERLI_KATEGORILER = {"Chiller", "MCC", "Kule", "TRDP", "Diger"}
CIHAZ_ADI_REGEX = re.compile(r"^[A-Z0-9]+(-[A-Z0-9]+)+$")

# Excel'deki "Sistem Rolu" secimini data_bridge.py'nin ic anahtarina cevirir.
# Hat1/Hat2 = sablondaki genel ad, Mas1/Mas2 = koddaki ic slot adi (Maslak'tan kalma, her
# lokasyonda Hat1->Mas1, Hat2->Mas2 olarak kullanilir — lokasyon adiyla ilgisi yoktur)
ROL_TO_ANAHTAR = {
    "Dis Hava Sicakligi":        "Dis_Hava_Sicakligi_C",
    "Chiller Set Sicakligi":     "Chiller_Set_Temp_C",
    "Hat1 Isitma Sicakligi":     "Mas1_Isitma_Temp",
    "Hat1 Sogutma Sicakligi":    "Mas1_Sogutma_Temp",
    "Hat1 Kazan Sicakligi":      "Mas1_Kazan_Temp",
    "Hat2 Isitma Sicakligi":     "Mas2_Isitma_Temp",
    "Hat2 Sogutma Sicakligi":    "Mas2_Sogutma_Temp",
    "Hat2 Kazan Sicakligi":      "Mas2_Kazan_Temp",
    "Absorpsiyon Chiller Durum": "_abs_dur",
    "Kullanilmayacak":           None,
}
for _i in range(1, 6):
    ROL_TO_ANAHTAR[f"Chiller-{_i} Durum"] = f"_ch_dur_{_i}"
    ROL_TO_ANAHTAR[f"Chiller-{_i} Yuzde"] = f"_ch_yuz_{_i}"
for _i in range(1, 4):
    ROL_TO_ANAHTAR[f"Kazan-{_i} Durum"] = f"_kaz_dur_{_i}"


def lokasyon_profili_oku(ws):
    """'Lokasyon Profili' sayfasindaki etiket->deger ciftlerini sozluge cevirir."""
    profil = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        etiket = row[0].value
        deger = row[1].value
        if not etiket or not isinstance(etiket, str):
            continue
        etiket = etiket.lstrip("* ").strip()
        if deger is not None and str(deger).strip() != "":
            profil[etiket] = str(deger).strip()
    return profil


ORNEK_CIHAZ_ADI = "MCC-1"
ORNEK_IP = "172.17.91.100"


def modbus_analizorler_oku(ws):
    """'Modbus Analizorler' sayfasindaki dolu satirlari [{ip,name,brand,kategori}] listesine cevirir."""
    analizorler = []
    uyarilar = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=6):
        cihaz_adi, ip, marka, kategori, notlar = (c.value for c in row[1:6])
        # Ornek satir SADECE 4. satirdadir (sablonun sabit konumu) — orada hala ornek
        # degerler varsa atla. Baska bir satirda aynen "MCC-1"/ayni IP girilmesi gercek
        # veri olabilir, onu atlama (Notlar metnine de guvenme, kullanici silmeyi unutabilir)
        if (row[0].row == 4 and str(cihaz_adi).strip() == ORNEK_CIHAZ_ADI
                and str(ip).strip() == ORNEK_IP):
            continue
        if not cihaz_adi and not ip:
            continue
        cihaz_adi = str(cihaz_adi).strip() if cihaz_adi else ""
        ip = str(ip).strip() if ip else ""
        marka = str(marka).strip().lower() if marka else ""
        kategori = str(kategori).strip() if kategori else ""

        if not cihaz_adi or not ip:
            uyarilar.append(f"Satir atlandi (Cihaz Adi veya IP eksik): {cihaz_adi or '?'} / {ip or '?'}")
            continue
        if not CIHAZ_ADI_REGEX.match(cihaz_adi):
            uyarilar.append(
                f"'{cihaz_adi}' formati BUYUK HARF+tire kuralina uymuyor "
                f"(orn. MCC-1, CHILLER-2) — kirilim raporlarinda eslesmeyebilir"
            )
        if kategori and kategori not in GECERLI_KATEGORILER:
            uyarilar.append(f"'{cihaz_adi}' icin gecersiz kategori: '{kategori}'")

        analizorler.append({"ip": ip, "name": cihaz_adi, "brand": marka or "janitza", "kategori": kategori})
    return analizorler, uyarilar


def bacnet_noktalar_oku(ws):
    """'BACnet Noktalar' sayfasindaki Point Name + Sistem Rolu ciftlerini okur."""
    eslesmeler = []  # [(point_name, anahtar)]
    uyarilar = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=9):
        point_name = row[6].value  # G: Point Name
        rol = row[8].value         # I: Sistem Rolu
        if not point_name or not str(point_name).strip():
            continue
        point_name = str(point_name).strip()
        rol = str(rol).strip() if rol else ""

        if not rol or rol == "Kullanilmayacak":
            continue
        if rol not in ROL_TO_ANAHTAR:
            uyarilar.append(f"'{point_name}' icin tanimsiz Sistem Rolu: '{rol}'")
            continue
        eslesmeler.append((point_name, ROL_TO_ANAHTAR[rol]))
    return eslesmeler, uyarilar


def hedefli_okuma_sablonu_uret(ws_bacnet) -> "Workbook":
    """'BACnet Noktalar' sayfasindaki ham satirlardan hedefli_okuma_sablonu_2.xlsx'i uretir
    (ahu_collector.py'nin gercekten okudugu format — Sistem Rolu sutunu olmadan, ayni 7 sutun)."""
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"
    basliklar = ["Gateway IP", "Network (DNET)", "MAC (DADR)", "Device Instance ID",
                 "Object Type", "Object Instance", "Point Name"]
    ws_out.append(basliklar)

    for row in ws_bacnet.iter_rows(min_row=4, max_row=ws_bacnet.max_row, max_col=9):
        degerler = [c.value for c in row[:7]]
        point_name = row[6].value
        if not point_name or not str(point_name).strip():
            continue
        ws_out.append(degerler)

    return wb_out


def data_bridge_uret(analizorler: list, bacnet_eslesmeler: list) -> str:
    """Mevcut data_bridge.py'yi sablon olarak kullanip analizor kategorilerini ve
    BACNET_MAP'i bu lokasyonun verisiyle degistirir."""
    if not os.path.exists(BRIDGE_SABLON):
        raise FileNotFoundError(f"Sablon dosya bulunamadi: {BRIDGE_SABLON}")
    with open(BRIDGE_SABLON, "r", encoding="utf-8") as f:
        kaynak = f.read()

    chiller_isimler = [a["name"] for a in analizorler if a.get("kategori") == "Chiller"]
    tum_isimler = [a["name"] for a in analizorler]

    def kume_blogu(isimler):
        if not isimler:
            return "set()"
        return "{\n" + "\n".join(f'    "{ad}",' for ad in isimler) + "\n}"

    kaynak, n1 = re.subn(
        r"CHILLER_ANALYZERS = \{.*?\}", "CHILLER_ANALYZERS = " + kume_blogu(chiller_isimler),
        kaynak, count=1, flags=re.DOTALL
    )
    kaynak, n2 = re.subn(
        r"ALL_ANALYZERS = \{.*?\}", "ALL_ANALYZERS = " + kume_blogu(tum_isimler),
        kaynak, count=1, flags=re.DOTALL
    )
    if n1 == 0 or n2 == 0:
        raise RuntimeError("Sablonda CHILLER_ANALYZERS/ALL_ANALYZERS bulunamadi — sablon degismis olabilir")

    if bacnet_eslesmeler:
        satirlar = "\n".join(f'    "{pn}":{" " * max(1, 28 - len(pn))}"{key}",' for pn, key in bacnet_eslesmeler)
        yeni_map = "BACNET_MAP = {\n" + satirlar + "\n}"
        kaynak, n3 = re.subn(r"BACNET_MAP = \{.*?\n\}", yeni_map, kaynak, count=1, flags=re.DOTALL)
        if n3 == 0:
            raise RuntimeError("Sablonda BACNET_MAP bulunamadi — sablon degismis olabilir")

    return kaynak


def _ahu_nokta_tipi(point_name: str) -> str:
    """ahu_collector.py _detect_point_type ile birebir ayni mantik (SAGLAMLASTIRMA F1).
    Start/Stop ve Basinc noktalari on kosul kapisi icin taninir; taninmayan ad artik
    Return VARSAYILMAZ, bos doner (cagiran satiri atlar)."""
    pn = point_name.lower().strip()
    if "start" in pn:
        return "Start/Stop"
    if ("bas" in pn and ("nc" in pn or "nç" in pn)) or "pressure" in pn:
        return "Basınç (Pa)"
    if pn.endswith(" set") or pn == "set":
        return "Set (°C)"
    if "fleme" in pn:
        return "SAT (°C)"
    if "so" in pn and "van" in pn:
        return "Cool Valve (%)"
    if "van" in pn:
        return "Heat Valve (%)"
    if "emi" in pn or "emü" in pn or "return" in pn or "dönü" in pn or "donu" in pn:
        return "Return (°C)"
    return ""  # taninmayan ad — konfige eklenmez


def _ahu_adi(point_name: str) -> str:
    """ahu_collector.py _ahu_adi ile birebir ayni mantik — ilk kelime AHU adidir."""
    parts = point_name.strip().split()
    return parts[0] if parts else "Unknown"


AHU_ORNEK_POINT_NAME = "Ahu-3 Ufleme"
AHU_ORNEK_GATEWAY_IP = "172.17.91.50"


def ahu_noktalari_oku(ws):
    """'AHU Noktalari' sayfasindan ahu_nokta_konfig.json icin satirlari uretir."""
    konfig = []
    uyarilar = []
    son_gateway, son_dnet, son_mac = "", "", ""
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=8):
        gateway_ip, dnet, mac_hex, dev_id, obj_type, obj_inst, point_name, mahal = (c.value for c in row)
        if (row[0].row == 4 and str(point_name).strip() == AHU_ORNEK_POINT_NAME
                and str(gateway_ip).strip() == AHU_ORNEK_GATEWAY_IP):
            continue
        if not point_name or not str(point_name).strip():
            continue
        point_name = str(point_name).strip()

        # Gateway/DNET/MAC bos ise bir onceki satirdan devral (Excel'de tekrar yazmamak icin)
        gateway_ip = str(gateway_ip).strip() if gateway_ip else son_gateway
        dnet_str = str(dnet).strip() if dnet is not None else son_dnet
        mac_hex = str(mac_hex).strip() if mac_hex else son_mac
        son_gateway, son_dnet, son_mac = gateway_ip, dnet_str, mac_hex

        _tip = _ahu_nokta_tipi(point_name)
        if not _tip:
            uyarilar.append(f"AHU noktasi TANINMADI, atlandi ('{point_name}') — "
                            "ad 'ufleme/emis/vana/set/start/basinc' anahtarlarindan birini icermeli")
            continue
        try:
            konfig.append({
                "ahu_adi":    _ahu_adi(point_name),
                "lokasyon":   str(mahal).strip() if mahal else "MAS-1",
                "nokta_tipi": _tip,
                "gateway_ip": gateway_ip,
                "dnet":       int(float(dnet_str)),
                "mac_hex":    mac_hex or "0x01",
                "obj_type":   int(float(obj_type)) if obj_type is not None else 0,
                "obj_inst":   int(float(obj_inst)) if obj_inst is not None else 0,
            })
        except (TypeError, ValueError) as e:
            uyarilar.append(f"AHU satiri atlandi ('{point_name}'): {e}")
    return konfig, uyarilar


def ahu_sat_kapasite_oku(ws):
    """'AHU SAT ve Kapasite' sayfasini okur — opsiyonel sayfa, bossa bos sonuc doner."""
    sat_limitleri = {}
    tasarim_kapasiteleri = {}
    uyarilar = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=5):
        mahal, ahu_adi, sogutma, isitma, kapasite = (c.value for c in row)
        if row[0].row == 4 and str(mahal).strip() == "MAS-1" and str(ahu_adi).strip() == "Ahu-1" \
                and sogutma == 18.0 and kapasite == 194.9:
            continue  # degismemis ornek satir
        if not mahal or not ahu_adi:
            continue
        mahal = str(mahal).strip()
        ahu_adi = str(ahu_adi).strip()

        if sogutma is not None or isitma is not None:
            limit = {}
            if sogutma is not None:
                limit["cooling"] = float(sogutma)
            if isitma is not None:
                limit["heating"] = float(isitma)
            sat_limitleri.setdefault(mahal, {})[ahu_adi] = limit

        if kapasite is not None:
            try:
                tasarim_kapasiteleri.setdefault(mahal, {})[ahu_adi] = float(kapasite)
            except (TypeError, ValueError):
                uyarilar.append(f"'{ahu_adi}' icin gecersiz kapasite degeri: '{kapasite}'")

    return sat_limitleri, tasarim_kapasiteleri, uyarilar


def data_collector_uret(analizorler: list) -> str:
    """Mevcut data_collector.py'yi sablon olarak kullanip ANALYZERS listesini degistirir."""
    if not os.path.exists(COLLECTOR_SABLON):
        raise FileNotFoundError(f"Sablon dosya bulunamadi: {COLLECTOR_SABLON}")
    with open(COLLECTOR_SABLON, "r", encoding="utf-8") as f:
        kaynak = f.read()

    satirlar = "\n".join(
        f'    {{"ip": "{a["ip"]}", "name": "{a["name"]}", "brand": "{a["brand"]}"}},'
        for a in analizorler
    )
    yeni_blok = f"ANALYZERS = [\n{satirlar}\n]"

    yeni_kaynak, sayi = re.subn(
        r"ANALYZERS = \[.*?\]", yeni_blok, kaynak, count=1, flags=re.DOTALL
    )
    if sayi == 0:
        raise RuntimeError("Sablonda ANALYZERS listesi bulunamadi — sablon degismis olabilir")
    return yeni_kaynak


SAT_BAND_VARSAYILAN = {
    "SAT_COOLING_MIN": 15.0, "SAT_COOLING_MAX": 18.0,
    "SAT_HEATING_MIN": 28.0, "SAT_HEATING_MAX": 31.0,   # SAGLAMLASTIRMA ile guncellendi
}

# SAGLAMLASTIRMA F5 — on kosul kapisi esikleri (her yeni lokasyona yazilir; on_kosul.py
# bu degerleri CONFIG'ten override alir, config_dogrula acilista bunlari denetler).
ON_KOSUL_VARSAYILAN = {
    "TARGET_AIR_DT_AHU_COOL": 6.0,     # AHU hava dT fallback (dinamik hedef esas: emis-16.5, 4-8 bandi)
    "TARGET_AIR_DT_AHU_HEAT": 10.0,
    "PRESSURE_RUN_MIN_PA": 20.0,       # ustu = fan calisiyor
    "PRESSURE_FAULT_MAX_PA": 1500.0,   # ustu = basinc sensoru arizasi
    "SENSOR_VALID_MIN_C": 1.0,         # ufleme/emis gecerli aralik alt
    "SENSOR_VALID_MAX_C": 50.0,        # gecerli aralik ust
    "MANUAL_SUPPRESS_HOURS": 24,       # elle OK = 24s susturma
    "DEBOUNCE_OKUMA": 2, "DUZELME_OKUMA": 3, "TAKILI_OKUMA": 12,
    "SKIP_ESKALASYON_GUN": 7,
}
SAT_BAND_PROFIL_ETIKETI = {
    "SAT_COOLING_MIN": "Sogutma Min Ufleme (SAT)",
    "SAT_COOLING_MAX": "Sogutma Max Ufleme (SAT)",
    "SAT_HEATING_MIN": "Isitma Min Ufleme (SAT)",
    "SAT_HEATING_MAX": "Isitma Max Ufleme (SAT)",
}

CHILLER_FCU_VARSAYILAN = {
    "chiller_adedi": None, "chiller_birim_kw": None,
    "chiller_t1_gidis": None, "chiller_t2_donus": None,
    "fcu_adedi": None, "fcu_birim_kw_ortalama": None,
    "fcu_esanjor_diversity": None, "tasarim_hava_dt": None,
}
CHILLER_FCU_PROFIL_ETIKETI = {
    "chiller_adedi": "Chiller Sayisi",
    "chiller_birim_kw": "Chiller Birim Kapasite (kW)",
    "chiller_t1_gidis": "Chiller Gidis Sicakligi (C)",
    "chiller_t2_donus": "Chiller Donus Sicakligi (C)",
    "fcu_adedi": "FCU Adedi",
    "fcu_birim_kw_ortalama": "FCU Birim Kapasite (kW)",
    "fcu_esanjor_diversity": "FCU Esanjor Diversity",
    "tasarim_hava_dt": "Tasarim Hava DT (C)",
}


def hvac_settings_uret(profil: dict) -> dict:
    """main_portal.py'nin kaynak kodunu DEGISTIRMEDEN, gercek mekanizmasi olan
    configs/hvac_settings.json'u uretir (main_portal.py acilista bu dosyayi okuyup
    CONFIG dict'ini override eder — load_settings_from_file())."""
    ayarlar = {}
    for anahtar, varsayilan in SAT_BAND_VARSAYILAN.items():
        etiket = SAT_BAND_PROFIL_ETIKETI[anahtar]
        deger_str = profil.get(etiket, "")
        try:
            ayarlar[anahtar] = float(deger_str) if deger_str else varsayilan
        except ValueError:
            ayarlar[anahtar] = varsayilan
    # SAGLAMLASTIRMA: on kosul kapisi esiklerini de yaz (main_portal fail-fast'i icin)
    ayarlar.update(ON_KOSUL_VARSAYILAN)
    return ayarlar


def chiller_fcu_ayarlari_uret(profil: dict) -> dict | None:
    """'Lokasyon Profili' sayfasindaki Chiller/FCU kapasite alanlarindan
    configs/chiller_fcu_ayarlari.json uretir. Hicbir alan girilmemisse None doner
    (dosya yazilmaz, ahu_collector.py kendi varsayilanini kullanir)."""
    ayarlar = {}
    for anahtar, etiket in CHILLER_FCU_PROFIL_ETIKETI.items():
        deger_str = profil.get(etiket, "")
        if not deger_str:
            continue
        try:
            ayarlar[anahtar] = int(float(deger_str)) if anahtar in ("chiller_adedi", "fcu_adedi") else float(deger_str)
        except ValueError:
            pass
    return ayarlar or None


def location_manager_uret(profil: dict, lokasyon_id: str, lokasyon_adi: str) -> str:
    """location_manager.py'yi uretir — app_portal.py'nin lokasyon kimligini ve tek/cift
    hat (heating_lines) davranisini belirleyen EN KRiTiK dosya. Hat Sayisi=2 ise dual-line
    (Mas1/Mas2 internal kod, isimle ilgisi yok — data_bridge.py BACNET_MAP ile ayni mantik),
    Hat Sayisi=1 ise tek hat (Isitma_Temp_C vb. genel sutunlar) modunu uretir."""
    hat_sayisi = str(profil.get("Hat Sayisi", "1")).strip()
    kisa_ad = lokasyon_id.capitalize()  # not: short_name su an kodda baska yerde okunmuyor, sadece bilgi amacli

    if hat_sayisi == "2":
        energy_schema = (
            '        "heating_lines": ["MAS-1", "MAS-2"],\n'
            '        "fields_per_line": ["heating_temp", "boiler_temp", "cooling_temp"],\n'
            '        "labels": {\n'
            '            "MAS-1": "Mas-1",\n'
            '            "MAS-2": "Mas-2"\n'
            '        }\n'
        )
        csv_columns = (
            '        "heating_temp_cols": [\n'
            '            "Mas1_Isitma_Temp", "Mas1_Kazan_Temp", "Mas1_Sogutma_Temp",\n'
            '            "Mas2_Isitma_Temp", "Mas2_Kazan_Temp", "Mas2_Sogutma_Temp"\n'
            '        ]\n'
        )
    else:
        energy_schema = (
            '        "heating_lines": [],\n'
            '        "fields_per_line": ["heating_temp", "boiler_temp", "cooling_temp"],\n'
            '        "labels": {}\n'
        )
        csv_columns = (
            '        "heating_temp_cols": [\n'
            '            "Isitma_Temp_C", "Kazan_Temp_C", "Sogutma_Temp_C"\n'
            '        ]\n'
        )

    return f'''# location_manager.py
# Tek lokasyon yonetimi — {lokasyon_adi}
# lokasyon_kurulum_otomasyon.py tarafindan uretildi.

from __future__ import annotations
import os
import json
import logging
from typing import Dict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

LOCATION_CONFIG = {{
    "id": "{lokasyon_id}",
    "name": "{lokasyon_adi}",
    "short_name": "{kisa_ad}",
    "is_master": False,
    "energy_schema": {{
{energy_schema}    }},
    "csv_columns": {{
{csv_columns}    }}
}}


class LocationManager:
    """Tek lokasyon yonetimi."""

    def get_active_location_id(self) -> str:
        return "{lokasyon_id}"

    def list_locations(self) -> list:
        return [LOCATION_CONFIG]

    def get_location_config(self, location_id: str = None) -> Dict:
        return LOCATION_CONFIG

    def get_location_dir(self, location_id: str = None) -> str:
        return BASE_DIR

    def get_data_path(self, filename: str, location_id: str = None) -> str:
        return os.path.join(BASE_DIR, filename)

    def ensure_locations_ready(self):
        pass


_manager = None

def get_manager() -> LocationManager:
    global _manager
    if _manager is None:
        _manager = LocationManager()
    return _manager
'''


def supabase_config_uret(lokasyon_id: str) -> dict:
    """merkez_config.json'daki paylasilan Supabase baglanti bilgisiyle config uretir."""
    merkez_config_path = os.path.join(BASE_DIR, "merkez", "configs", "merkez_config.json")
    supabase_url, supabase_key = "BURAYA_SUPABASE_URL", "BURAYA_SUPABASE_KEY"
    if os.path.exists(merkez_config_path):
        with open(merkez_config_path, "r", encoding="utf-8") as f:
            mc = json.load(f)
        supabase_url = mc.get("supabase_url", supabase_url)
        supabase_key = mc.get("supabase_key", supabase_key)
    return {
        "lokasyon_id": lokasyon_id,
        "supabase_url": supabase_url,
        "supabase_key": supabase_key,
    }


def lisans_kaydi_yayinla(lokasyon_id: str, makine_id: str, makine_adi: str) -> dict:
    """Supabase 'lisanslar' tablosuna aktif lisans kaydi ekler/gunceller (upsert,
    makine_id+lokasyon_id ikilisi uzerinden). service_role_key gerekir."""
    import urllib.request

    secret_path = os.path.join(BASE_DIR, "hvac", "deneme", "supabase_secret.json")
    merkez_config_path = os.path.join(BASE_DIR, "merkez", "configs", "merkez_config.json")
    with open(secret_path, "r", encoding="utf-8") as f:
        service_key = json.load(f)["service_role_key"]
    with open(merkez_config_path, "r", encoding="utf-8") as f:
        supabase_url = json.load(f)["supabase_url"]

    payload = json.dumps({
        "lokasyon_id": lokasyon_id,
        "makine_id": makine_id,
        "makine_adi": makine_adi,
        "aktif": True,
    }).encode("utf-8")

    req = urllib.request.Request(
        supabase_url + "/rest/v1/lisanslar?on_conflict=makine_id,lokasyon_id",
        data=payload,
        headers={
            "apikey": service_key,
            "Authorization": "Bearer " + service_key,
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=representation",
        },
        method="POST",
    )
    resp = urllib.request.urlopen(req)
    return json.loads(resp.read().decode())[0]


def main():
    if len(sys.argv) < 2:
        print("Kullanim: python lokasyon_kurulum_otomasyon.py <doldurulmus_sablon.xlsx>")
        sys.exit(1)

    excel_yolu = sys.argv[1]
    if not os.path.exists(excel_yolu):
        print(f"HATA: dosya bulunamadi: {excel_yolu}")
        sys.exit(1)

    wb = load_workbook(excel_yolu, data_only=True)
    profil = lokasyon_profili_oku(wb["Lokasyon Profili"])
    analizorler, uyarilar = modbus_analizorler_oku(wb["Modbus Analizorler"])
    bacnet_eslesmeler, bacnet_uyarilar = bacnet_noktalar_oku(wb["BACnet Noktalar"])
    uyarilar += bacnet_uyarilar
    ahu_konfig, ahu_uyarilar = ahu_noktalari_oku(wb["AHU Noktalari"])
    uyarilar += ahu_uyarilar
    sat_limitleri, tasarim_kapasiteleri, sat_uyarilar = ahu_sat_kapasite_oku(wb["AHU SAT ve Kapasite"])
    uyarilar += sat_uyarilar

    lokasyon_id = profil.get("Lokasyon ID", "").strip()
    lokasyon_adi = profil.get("Lokasyon Adi", "").strip()

    eksik_zorunlu = [k for k in ("Lokasyon ID", "Lokasyon Adi") if not profil.get(k)]
    if eksik_zorunlu:
        print(f"HATA: zorunlu alanlar bos: {', '.join(eksik_zorunlu)}")
        sys.exit(1)

    cikti_dir = os.path.join(BASE_DIR, "lokasyonlar", lokasyon_id)
    os.makedirs(cikti_dir, exist_ok=True)

    if analizorler:
        collector_kaynak = data_collector_uret(analizorler)
        collector_yolu = os.path.join(cikti_dir, "data_collector.py")
        with open(collector_yolu, "w", encoding="utf-8") as f:
            f.write(collector_kaynak)
        print(f"OLUSTURULDU: {collector_yolu}  ({len(analizorler)} analizor)")
    else:
        print("ATLANDI: data_collector.py — Modbus Analizorler sayfasi bos")

    if analizorler or bacnet_eslesmeler:
        bridge_kaynak = data_bridge_uret(analizorler, bacnet_eslesmeler)
        bridge_yolu = os.path.join(cikti_dir, "data_bridge.py")
        with open(bridge_yolu, "w", encoding="utf-8") as f:
            f.write(bridge_kaynak)
        print(f"OLUSTURULDU: {bridge_yolu}  ({len(bacnet_eslesmeler)} BACnet eslesmesi)")
    else:
        print("ATLANDI: data_bridge.py — Modbus ve BACnet sayfalari bos")

    if bacnet_eslesmeler:
        hedefli_wb = hedefli_okuma_sablonu_uret(wb["BACnet Noktalar"])
        hedefli_yolu = os.path.join(cikti_dir, "configs", "hedefli_okuma_sablonu_2.xlsx")
        os.makedirs(os.path.dirname(hedefli_yolu), exist_ok=True)
        hedefli_wb.save(hedefli_yolu)
        print(f"OLUSTURULDU: {hedefli_yolu}")
    else:
        print("ATLANDI: hedefli_okuma_sablonu_2.xlsx — BACnet Noktalar sayfasi bos")

    loc_mgr_kaynak = location_manager_uret(profil, lokasyon_id, lokasyon_adi)
    loc_mgr_yolu = os.path.join(cikti_dir, "location_manager.py")
    with open(loc_mgr_yolu, "w", encoding="utf-8") as f:
        f.write(loc_mgr_kaynak)
    print(f"OLUSTURULDU: {loc_mgr_yolu}  (Hat Sayisi: {profil.get('Hat Sayisi', '1')})")

    hvac_settings = hvac_settings_uret(profil)
    settings_dir = os.path.join(cikti_dir, "configs")
    os.makedirs(settings_dir, exist_ok=True)
    settings_yolu = os.path.join(settings_dir, "hvac_settings.json")
    with open(settings_yolu, "w", encoding="utf-8") as f:
        json.dump(hvac_settings, f, ensure_ascii=False, indent=2)
    print(f"OLUSTURULDU: {settings_yolu}  ({hvac_settings})")

    chiller_fcu = chiller_fcu_ayarlari_uret(profil)
    if chiller_fcu:
        cf_yolu = os.path.join(settings_dir, "chiller_fcu_ayarlari.json")
        with open(cf_yolu, "w", encoding="utf-8") as f:
            json.dump(chiller_fcu, f, ensure_ascii=False, indent=2)
        print(f"OLUSTURULDU: {cf_yolu}")
    else:
        print("ATLANDI: chiller_fcu_ayarlari.json — Chiller/FCU Kapasite Bilgisi alanlari bos (opsiyonel)")

    config = supabase_config_uret(lokasyon_id)
    config_yolu = os.path.join(cikti_dir, "supabase_config.json")
    with open(config_yolu, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    print(f"OLUSTURULDU: {config_yolu}")

    makine_id = profil.get("Makine ID", "").strip()
    if makine_id:
        try:
            lisans_kaydi_yayinla(lokasyon_id, makine_id, f"{lokasyon_adi} PC")
            print(f"SUPABASE: lisanslar tablosuna kayit eklendi/guncellendi (makine_id={makine_id})")
        except Exception as e:
            print(f"UYARI: lisans kaydi olusturulamadi: {e}")
    else:
        print("ATLANDI: lisanslar tablosu kaydi — Makine ID alani bos")

    if ahu_konfig:
        ahu_yolu = os.path.join(cikti_dir, "ahu_nokta_konfig.json")
        with open(ahu_yolu, "w", encoding="utf-8") as f:
            json.dump(ahu_konfig, f, ensure_ascii=False, indent=2)
        print(f"OLUSTURULDU: {ahu_yolu}  ({len(ahu_konfig)} AHU noktasi)")
    else:
        print("ATLANDI: ahu_nokta_konfig.json — AHU Noktalari sayfasi bos")

    if sat_limitleri:
        os.makedirs(os.path.join(cikti_dir, "configs"), exist_ok=True)
        sat_yolu = os.path.join(cikti_dir, "configs", "ahu_sat_limitleri.json")
        with open(sat_yolu, "w", encoding="utf-8") as f:
            json.dump(sat_limitleri, f, ensure_ascii=False, indent=2)
        print(f"OLUSTURULDU: {sat_yolu}")
    else:
        print("ATLANDI: ahu_sat_limitleri.json — AHU SAT ve Kapasite sayfasi bos (opsiyonel)")

    if tasarim_kapasiteleri:
        os.makedirs(os.path.join(cikti_dir, "configs"), exist_ok=True)
        kap_yolu = os.path.join(cikti_dir, "configs", "ahu_tasarim_kapasiteleri.json")
        with open(kap_yolu, "w", encoding="utf-8") as f:
            json.dump(tasarim_kapasiteleri, f, ensure_ascii=False, indent=2)
        print(f"OLUSTURULDU: {kap_yolu}")
    else:
        print("ATLANDI: ahu_tasarim_kapasiteleri.json — AHU SAT ve Kapasite sayfasi bos (opsiyonel)")

    zip_yolu = shutil.make_archive(cikti_dir, "zip", cikti_dir)
    print(f"OLUSTURULDU: {zip_yolu}  (kurulum zip paketi)")

    print()
    print(f"--- Lokasyon: {lokasyon_id} ({lokasyon_adi}) ---")
    if uyarilar:
        print(f"\n{len(uyarilar)} UYARI bulundu:")
        for u in uyarilar:
            print(f"  ! {u}")
    else:
        print("\nUyari yok.")

    print(
        "\nMANUEL YAPILMASI GEREKENLER (henuz otomatik degil):\n"
        "  - Supabase 'lokasyonlar' tablosuna kayit\n"
        "  - app_merkez.py guncellemesi (HASTANELER dict'ine yeni lokasyon eklenir)"
    )


if __name__ == "__main__":
    main()
